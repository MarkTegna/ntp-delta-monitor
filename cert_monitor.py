#!/usr/bin/env python3
"""
Certificate Expiry Monitor - A Windows-based HTTPS certificate monitoring program
Checks SSL/TLS certificate expiration dates for a list of servers

Author: Mark Oldham
"""

import argparse
import configparser
import csv
import logging
import os
import smtplib
import socket
import ssl
import struct
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Optional, List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Program version
VERSION = "0.0.1"
PROGRAM_NAME = "Certificate Expiry Monitor"


@dataclass
class CertResult:
    """Data model for certificate check result"""
    timestamp_utc: datetime
    server: str
    port: int
    short_name: str
    server_ip: Optional[str]
    status: str  # OK, NO_RESPONSE, SSL_ERROR, ERROR
    cert_expiry: Optional[str]  # YYYY-MM-DD
    days_remaining: Optional[int]
    cert_subject: Optional[str]
    cert_issuer: Optional[str]
    error_message: Optional[str] = None


@dataclass
class Config:
    """Configuration data model"""
    servers_file: Path
    output_file: Optional[Path]
    parallel_limit: int
    timeout: int
    warning_days: int  # Days before expiry to flag as warning
    critical_days: int  # Days before expiry to flag as critical
    verbose: bool


def _parse_cert_expiry_from_der(der_bytes: bytes) -> Optional[str]:
    """
    Parse certificate expiry date from DER-encoded certificate bytes.
    Walks the ASN.1 structure to find the Validity sequence and extract notAfter.

    Returns expiry as 'YYYY-MM-DD' string, or None if parsing fails.
    """
    try:
        def _read_tag_len(data, offset):
            if offset >= len(data):
                return None, 0, 0
            tag = data[offset]
            if offset + 1 >= len(data):
                return tag, 0, 1
            length_byte = data[offset + 1]
            if length_byte < 0x80:
                return tag, length_byte, 2
            num_bytes = length_byte & 0x7f
            if offset + 2 + num_bytes > len(data):
                return tag, 0, 2
            length = int.from_bytes(data[offset + 2:offset + 2 + num_bytes], 'big')
            return tag, length, 2 + num_bytes

        def _parse_time(data, offset):
            tag, length, hdr = _read_tag_len(data, offset)
            if tag not in (0x17, 0x18) or length == 0:
                return None
            time_str = data[offset + hdr:offset + hdr + length].decode('ascii')
            if tag == 0x17:  # UTCTime
                year = int(time_str[:2])
                year += 2000 if year < 50 else 1900
                return f"{year:04d}-{int(time_str[2:4]):02d}-{int(time_str[4:6]):02d}"
            else:  # GeneralizedTime
                return f"{int(time_str[:4]):04d}-{int(time_str[4:6]):02d}-{int(time_str[6:8]):02d}"

        tag, cert_len, hdr = _read_tag_len(der_bytes, 0)
        if tag != 0x30:
            return None
        tbs_offset = hdr
        tag, tbs_len, tbs_hdr = _read_tag_len(der_bytes, tbs_offset)
        if tag != 0x30:
            return None

        pos = tbs_offset + tbs_hdr
        tbs_end = tbs_offset + tbs_hdr + tbs_len
        fields_skipped = 0
        while pos < tbs_end and fields_skipped < 4:
            tag, length, field_hdr = _read_tag_len(der_bytes, pos)
            if tag is None:
                break
            pos += field_hdr + length
            fields_skipped += 1

        if pos >= tbs_end:
            return None
        tag, val_len, val_hdr = _read_tag_len(der_bytes, pos)
        if tag != 0x30:
            return None

        nb_offset = pos + val_hdr
        nb_tag, nb_len, nb_hdr = _read_tag_len(der_bytes, nb_offset)
        na_offset = nb_offset + nb_hdr + nb_len
        return _parse_time(der_bytes, na_offset)
    except Exception:
        return None


def _extract_cert_field(der_bytes: bytes, field_index: int) -> Optional[str]:
    """Extract subject (field_index=5) or issuer (field_index=3) CN from DER cert."""
    try:
        def _read_tag_len(data, offset):
            if offset >= len(data):
                return None, 0, 0
            tag = data[offset]
            if offset + 1 >= len(data):
                return tag, 0, 1
            length_byte = data[offset + 1]
            if length_byte < 0x80:
                return tag, length_byte, 2
            num_bytes = length_byte & 0x7f
            if offset + 2 + num_bytes > len(data):
                return tag, 0, 2
            length = int.from_bytes(data[offset + 2:offset + 2 + num_bytes], 'big')
            return tag, length, 2 + num_bytes

        tag, cert_len, hdr = _read_tag_len(der_bytes, 0)
        if tag != 0x30:
            return None
        tbs_offset = hdr
        tag, tbs_len, tbs_hdr = _read_tag_len(der_bytes, tbs_offset)
        if tag != 0x30:
            return None

        pos = tbs_offset + tbs_hdr
        tbs_end = tbs_offset + tbs_hdr + tbs_len
        fields_skipped = 0
        while pos < tbs_end and fields_skipped < field_index:
            tag, length, field_hdr = _read_tag_len(der_bytes, pos)
            if tag is None:
                break
            pos += field_hdr + length
            fields_skipped += 1

        if pos >= tbs_end:
            return None

        # Now at the target field - try to extract readable strings
        tag, length, field_hdr = _read_tag_len(der_bytes, pos)
        field_data = der_bytes[pos + field_hdr:pos + field_hdr + length]

        # Find printable string fields (tags 0x0c=UTF8, 0x13=PrintableString, 0x16=IA5String)
        parts = []
        i = 0
        while i < len(field_data) - 2:
            t = field_data[i]
            if t in (0x0c, 0x13, 0x16):
                slen = field_data[i + 1]
                if slen < 0x80 and i + 2 + slen <= len(field_data):
                    try:
                        s = field_data[i + 2:i + 2 + slen].decode('utf-8')
                        if len(s) > 1:
                            parts.append(s)
                    except UnicodeDecodeError:
                        pass
                    i += 2 + slen
                    continue
            i += 1
        return ', '.join(parts) if parts else None
    except Exception:
        return None


def check_certificate(server: str, port: int = 443, timeout: int = 10,
                      short_name: str = '') -> CertResult:
    """
    Connect to server via HTTPS and extract certificate expiration info.

    Args:
        server: Hostname or IP address
        port: HTTPS port (default 443)
        timeout: Connection timeout in seconds
        short_name: Display name for the server

    Returns:
        CertResult with certificate details
    """
    logger = logging.getLogger(__name__)
    query_timestamp = datetime.now(timezone.utc)

    # Resolve IP
    resolved_ip = None
    try:
        resolved_ip = socket.gethostbyname(server)
    except socket.gaierror:
        pass

    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        ssl_sock = ctx.wrap_socket(sock, server_hostname=server)
        ssl_sock.connect((server, port))

        cert_der = ssl_sock.getpeercert(binary_form=True)
        ssl_sock.close()

        if not cert_der:
            logger.debug(f"Cert check {server}:{port}: no certificate returned")
            return CertResult(
                timestamp_utc=query_timestamp, server=server, port=port,
                short_name=short_name or server, server_ip=resolved_ip,
                status='SSL_ERROR', cert_expiry=None, days_remaining=None,
                cert_subject=None, cert_issuer=None,
                error_message='No certificate returned')

        # Parse expiry
        cert_expiry = _parse_cert_expiry_from_der(cert_der)
        days_remaining = None
        if cert_expiry:
            try:
                expiry_date = datetime.strptime(cert_expiry, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                days_remaining = (expiry_date - datetime.now(timezone.utc)).days
            except ValueError:
                pass

        # Extract subject and issuer
        cert_subject = _extract_cert_field(cert_der, 5)
        cert_issuer = _extract_cert_field(cert_der, 3)

        logger.debug(f"Cert check {server}:{port}: expiry={cert_expiry}, days={days_remaining}")

        return CertResult(
            timestamp_utc=query_timestamp, server=server, port=port,
            short_name=short_name or server, server_ip=resolved_ip,
            status='OK', cert_expiry=cert_expiry, days_remaining=days_remaining,
            cert_subject=cert_subject, cert_issuer=cert_issuer)

    except socket.timeout:
        logger.debug(f"Cert check {server}:{port}: timed out")
        return CertResult(
            timestamp_utc=query_timestamp, server=server, port=port,
            short_name=short_name or server, server_ip=resolved_ip,
            status='NO_RESPONSE', cert_expiry=None, days_remaining=None,
            cert_subject=None, cert_issuer=None,
            error_message=f'Connection timed out after {timeout}s')

    except ConnectionRefusedError:
        logger.debug(f"Cert check {server}:{port}: connection refused")
        return CertResult(
            timestamp_utc=query_timestamp, server=server, port=port,
            short_name=short_name or server, server_ip=resolved_ip,
            status='NO_RESPONSE', cert_expiry=None, days_remaining=None,
            cert_subject=None, cert_issuer=None,
            error_message='Connection refused')

    except ssl.SSLError as e:
        logger.debug(f"Cert check {server}:{port}: SSL error: {e}")
        return CertResult(
            timestamp_utc=query_timestamp, server=server, port=port,
            short_name=short_name or server, server_ip=resolved_ip,
            status='SSL_ERROR', cert_expiry=None, days_remaining=None,
            cert_subject=None, cert_issuer=None,
            error_message=f'SSL error: {e}')

    except Exception as e:
        logger.debug(f"Cert check {server}:{port}: error: {e}")
        return CertResult(
            timestamp_utc=query_timestamp, server=server, port=port,
            short_name=short_name or server, server_ip=resolved_ip,
            status='ERROR', cert_expiry=None, days_remaining=None,
            cert_subject=None, cert_issuer=None,
            error_message=f'{e}')


def parse_servers_csv(file_path: Path) -> List[tuple[str, int, str]]:
    """
    Parse CSV file with server list. Expects 'server' column, optional 'port' and 'short_name'.

    Returns list of (server, port, short_name) tuples.
    """
    logger = logging.getLogger(__name__)
    servers = []

    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames or 'server' not in reader.fieldnames:
                logger.error(f"CSV missing 'server' column. Found: {reader.fieldnames}")
                sys.exit(1)

            has_port = 'port' in reader.fieldnames
            has_name = 'short_name' in reader.fieldnames

            for row_num, row in enumerate(reader, 2):
                server = row.get('server', '').strip()
                if not server:
                    continue

                port = 443
                if has_port:
                    port_str = row.get('port', '').strip()
                    if port_str:
                        try:
                            port = int(port_str)
                        except ValueError:
                            port = 443

                short_name = ''
                if has_name:
                    short_name = row.get('short_name', '').strip()
                if not short_name:
                    short_name = server

                servers.append((server, port, short_name))
                logger.debug(f"Row {row_num}: {server}:{port} ({short_name})")

        logger.info(f"Parsed {len(servers)} servers from {file_path}")
        return servers

    except Exception as e:
        logger.error(f"Error reading CSV {file_path}: {e}")
        sys.exit(1)


def process_servers_parallel(servers: List[tuple[str, int, str]], config: Config) -> List[CertResult]:
    """Process all servers concurrently."""
    logger = logging.getLogger(__name__)
    results = []

    if not servers:
        return results

    max_workers = config.parallel_limit if config.parallel_limit > 0 else 10
    logger.info(f"Processing {len(servers)} servers with {max_workers} workers")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_server = {}
        for server, port, short_name in servers:
            future = executor.submit(check_certificate, server, port, config.timeout, short_name)
            future_to_server[future] = server

        for future in as_completed(future_to_server):
            server = future_to_server[future]
            try:
                result = future.result()
                results.append(result)
            except Exception as e:
                logger.error(f"Unexpected error for {server}: {e}")
                results.append(CertResult(
                    timestamp_utc=datetime.now(timezone.utc), server=server, port=443,
                    short_name=server, server_ip=None, status='ERROR',
                    cert_expiry=None, days_remaining=None,
                    cert_subject=None, cert_issuer=None,
                    error_message=f'Unexpected: {e}'))

    # Sort: errors first, then by days_remaining ascending (soonest expiry first)
    def sort_key(r):
        if r.status != 'OK':
            return (0, 0)
        if r.days_remaining is not None:
            return (1, r.days_remaining)
        return (2, 0)

    results.sort(key=sort_key)
    return results


def _sanitize_xlsx(value: str) -> str:
    """Remove control characters that openpyxl rejects."""
    import re
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)


def write_xlsx_report(results: List[CertResult], output_path: Path, config: Config) -> None:
    """Write results to XLSX with formatting."""
    logger = logging.getLogger(__name__)

    try:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Certificate Expiry Report"

        headers = [
            'Timestamp (UTC)',
            'Server',
            'Port',
            'Short Name',
            'Server IP',
            'Cert Expiry',
            'Days Remaining',
            'Subject',
            'Issuer',
            'Status',
            'Error Message'
        ]

        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows
        for row_num, result in enumerate(results, 2):
            row_data = [
                result.timestamp_utc.isoformat(),
                result.server,
                result.port,
                result.short_name,
                result.server_ip or '',
                result.cert_expiry or '',
                result.days_remaining if result.days_remaining is not None else '',
                _sanitize_xlsx(result.cert_subject or ''),
                _sanitize_xlsx(result.cert_issuer or ''),
                result.status,
                result.error_message or ''
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

            # Status cell formatting (column 10)
            status_cell = ws.cell(row=row_num, column=10)
            if result.status == 'OK':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(bold=True)

            # Days remaining formatting (column 7)
            days_cell = ws.cell(row=row_num, column=7)
            if result.days_remaining is not None:
                if result.days_remaining <= config.critical_days:
                    # Critical - red
                    days_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    days_cell.font = Font(color="FFFFFF", bold=True)
                elif result.days_remaining <= config.warning_days:
                    # Warning - yellow
                    days_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    days_cell.font = Font(bold=True)
                else:
                    # OK - green
                    days_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    days_cell.font = Font(bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            try:
                col_letter = column[0].column_letter
            except AttributeError:
                continue
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter') and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        workbook.save(output_path)
        logger.info(f"Wrote {len(results)} results to {output_path}")

    except Exception as e:
        logger.error(f"Failed to write XLSX: {e}")
        raise


def format_summary(results: List[CertResult], config: Config) -> str:
    """Generate summary text."""
    total = len(results)
    ok = sum(1 for r in results if r.status == 'OK')
    failed = total - ok
    expired = sum(1 for r in results if r.days_remaining is not None and r.days_remaining <= 0)
    critical = sum(1 for r in results if r.days_remaining is not None and 0 < r.days_remaining <= config.critical_days)
    warning = sum(1 for r in results if r.days_remaining is not None and config.critical_days < r.days_remaining <= config.warning_days)

    lines = [
        "=" * 60,
        "CERTIFICATE EXPIRY SUMMARY",
        "=" * 60,
        f"Hostname: {socket.gethostname()}",
        f"Program: {PROGRAM_NAME} v{VERSION}",
        "",
        f"Total servers checked: {total}",
        f"Successful checks: {ok}",
        f"Failed checks: {failed}",
        "",
        f"Expired certificates: {expired}",
        f"Critical (< {config.critical_days} days): {critical}",
        f"Warning (< {config.warning_days} days): {warning}",
    ]

    # List problem certs
    problems = [r for r in results if r.days_remaining is not None and r.days_remaining <= config.warning_days]
    if problems:
        lines.append("")
        lines.append("Certificates requiring attention:")
        for r in sorted(problems, key=lambda x: x.days_remaining or 0):
            status = "EXPIRED" if r.days_remaining <= 0 else f"{r.days_remaining} days"
            lines.append(f"  {r.short_name} ({r.server}:{r.port}): {status} - expires {r.cert_expiry}")

    # List failed checks
    failures = [r for r in results if r.status != 'OK']
    if failures:
        lines.append("")
        lines.append("Failed checks:")
        for r in failures:
            lines.append(f"  {r.short_name} ({r.server}:{r.port}): {r.status} - {r.error_message}")

    lines.append("=" * 60)
    return "\n".join(lines)


def send_email_notification(summary_text: str, xlsx_path: Path, results: List[CertResult],
                            config: Config, ini_config: dict) -> None:
    """Send email notification with results."""
    logger = logging.getLogger(__name__)

    if not ini_config.get('send_email', False):
        return

    try:
        expired = sum(1 for r in results if r.days_remaining is not None and r.days_remaining <= 0)
        critical = sum(1 for r in results if r.days_remaining is not None and 0 < r.days_remaining <= config.critical_days)
        failed = sum(1 for r in results if r.status != 'OK')

        has_error = expired > 0 or critical > 0 or failed > 0
        prefix = "CERT ERROR" if has_error else "CERT REPORT"

        issues = []
        if expired > 0:
            issues.append(f"{expired} expired")
        if critical > 0:
            issues.append(f"{critical} critical")
        if failed > 0:
            issues.append(f"{failed} unreachable")

        status_part = ', '.join(issues) if issues else "all certificates valid"
        subject = f"{prefix} - {status_part}"

        msg = MIMEMultipart()
        msg['From'] = ini_config.get('from_email', 'cert-monitor@tgna.tegna.com')
        msg['To'] = ini_config.get('to_email', 'moldham@tegna.com')
        msg['Subject'] = subject

        if has_error:
            msg['X-Priority'] = '1'
            msg['X-MSMail-Priority'] = 'High'
            msg['Importance'] = 'High'

        msg.attach(MIMEText(summary_text, 'plain'))

        if xlsx_path.exists():
            with open(xlsx_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename= {xlsx_path.name}')
                msg.attach(part)

        smtp_server = ini_config.get('smtp_server', 'relay.tgna.tegna.com')
        smtp_port = ini_config.get('smtp_port', 25)

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            if ini_config.get('smtp_use_tls', False):
                server.starttls()
            username = ini_config.get('smtp_username', '')
            password = ini_config.get('smtp_password', '')
            if username and password:
                server.login(username, password)
            server.send_message(msg)
            logger.info(f"Email sent: {subject}")

    except Exception as e:
        logger.error(f"Failed to send email: {e}")


def load_configuration(config_file: str = "cert_monitor.ini") -> dict:
    """Load configuration from INI file with defaults."""
    logger = logging.getLogger(__name__)

    defaults = {
        'default_parallel_limit': 10,
        'default_timeout': 10,
        'warning_days': 30,
        'critical_days': 7,
        'output_directory': '.\\Reports',
        'send_email': False,
        'smtp_server': '',
        'smtp_port': 25,
        'smtp_use_tls': False,
        'smtp_username': '',
        'smtp_password': '',
        'from_email': '',
        'to_email': '',
    }

    config = configparser.ConfigParser()
    try:
        if Path(config_file).exists():
            config.read(config_file)
            logger.debug(f"Loaded config from {config_file}")

            if 'monitor_settings' in config:
                s = config['monitor_settings']
                defaults['default_parallel_limit'] = s.getint('default_parallel_limit', defaults['default_parallel_limit'])
                defaults['default_timeout'] = s.getint('default_timeout', defaults['default_timeout'])
                defaults['warning_days'] = s.getint('warning_days', defaults['warning_days'])
                defaults['critical_days'] = s.getint('critical_days', defaults['critical_days'])
                defaults['output_directory'] = s.get('output_directory', defaults['output_directory'])

            if 'email_settings' in config:
                s = config['email_settings']
                defaults['send_email'] = s.getboolean('send_email', defaults['send_email'])
                defaults['smtp_server'] = s.get('smtp_server', defaults['smtp_server'])
                defaults['smtp_port'] = s.getint('smtp_port', defaults['smtp_port'])
                defaults['smtp_use_tls'] = s.getboolean('smtp_use_tls', defaults['smtp_use_tls'])
                defaults['smtp_username'] = s.get('smtp_username', defaults['smtp_username'])
                defaults['smtp_password'] = s.get('smtp_password', defaults['smtp_password'])
                defaults['from_email'] = s.get('from_email', defaults['from_email'])
                defaults['to_email'] = s.get('to_email', defaults['to_email'])
        else:
            logger.debug(f"Config file {config_file} not found, using defaults")
    except Exception as e:
        logger.warning(f"Error reading config: {e}")

    return defaults


def setup_logging(verbose: bool = False) -> None:
    """Configure logging."""
    log_level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=log_level, format='%(asctime)s - %(levelname)s - %(message)s',
                        handlers=[logging.StreamHandler(sys.stdout)])


def parse_arguments() -> Config:
    """Parse CLI arguments."""
    ini_config = load_configuration()

    parser = argparse.ArgumentParser(
        description='Certificate Expiry Monitor - Check HTTPS certificate expiration dates',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Examples:
  Basic usage:
    %(prog)s -s servers.csv

  Custom output and timeout:
    %(prog)s -s servers.csv -o report.xlsx -t 5

  With warning/critical thresholds:
    %(prog)s -s servers.csv --warning-days 60 --critical-days 14

CSV Format:
  Required column: server
  Optional columns: port (default 443), short_name

  Example:
    server,port,short_name
    www.example.com,443,Example
    internal.corp.com,8443,Internal App
    10.0.0.1,,Web Server
        """)

    parser.add_argument('-s', '--servers-file', type=Path, required=True, metavar='FILE',
                        help='Path to CSV file with server list (required)')
    parser.add_argument('-o', '--output-file', type=Path, metavar='FILE',
                        help='Output XLSX file path (default: auto-generated with timestamp)')
    parser.add_argument('-p', '--parallel-limit', type=int,
                        default=ini_config['default_parallel_limit'], metavar='N',
                        help=f'Max concurrent checks [default: {ini_config["default_parallel_limit"]}]')
    parser.add_argument('-t', '--timeout', type=int,
                        default=ini_config['default_timeout'], metavar='SECONDS',
                        help=f'Connection timeout [default: {ini_config["default_timeout"]}]')
    parser.add_argument('--warning-days', type=int,
                        default=ini_config['warning_days'], metavar='N',
                        help=f'Days before expiry to flag as warning (yellow) [default: {ini_config["warning_days"]}]')
    parser.add_argument('--critical-days', type=int,
                        default=ini_config['critical_days'], metavar='N',
                        help=f'Days before expiry to flag as critical (red) [default: {ini_config["critical_days"]}]')
    parser.add_argument('-v', '--verbose', action='store_true',
                        help='Enable verbose logging')
    parser.add_argument('--version', action='version',
                        version=f'{PROGRAM_NAME} {VERSION}')

    args = parser.parse_args()

    if not args.servers_file.exists():
        parser.error(f"Servers file not found: {args.servers_file}")

    return Config(
        servers_file=args.servers_file,
        output_file=args.output_file,
        parallel_limit=args.parallel_limit,
        timeout=args.timeout,
        warning_days=args.warning_days,
        critical_days=args.critical_days,
        verbose=args.verbose
    )


def main():
    """Main entry point."""
    try:
        ini_config = load_configuration()
        config = parse_arguments()
        setup_logging(config.verbose)

        logger = logging.getLogger(__name__)
        logger.info(f"{PROGRAM_NAME} v{VERSION} starting...")
        logger.info(f"Servers file: {config.servers_file}")
        logger.info(f"Warning threshold: {config.warning_days} days")
        logger.info(f"Critical threshold: {config.critical_days} days")

        # Parse server list
        servers = parse_servers_csv(config.servers_file)
        if not servers:
            logger.error("No servers to check")
            return 1

        # Process servers
        results = process_servers_parallel(servers, config)

        # Determine output path
        if config.output_file:
            output_path = config.output_file
        else:
            output_dir = Path(ini_config.get('output_directory', '.\\Reports'))
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"cert_monitor_report_{timestamp}.xlsx"

        # Write report
        write_xlsx_report(results, output_path, config)

        # Generate and display summary
        summary = format_summary(results, config)
        print(summary)

        # Write summary file
        summary_path = output_path.with_suffix('.txt')
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(summary + '\n')
        logger.info(f"Summary written to {summary_path}")

        # Send email
        send_email_notification(summary, output_path, results, config, ini_config)

        # Exit code
        has_errors = any(r.status != 'OK' for r in results)
        has_expired = any(r.days_remaining is not None and r.days_remaining <= 0 for r in results)

        if has_errors or has_expired:
            logger.warning("Completed with issues")
            return 1

        logger.info("Completed successfully")
        return 0

    except KeyboardInterrupt:
        print("\nCancelled by user", file=sys.stderr)
        return 1
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
