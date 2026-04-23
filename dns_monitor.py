#!/usr/bin/env python3
"""
DNS Response Monitor - A Windows-based DNS server monitoring program
Tests DNS response times across domain controllers and additional servers

Author: Mark Oldham
"""

import argparse
import configparser
import csv
import logging
import os
import smtplib
import socket
import statistics
import struct
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Optional, List
import random
import time

import dns.resolver
import dns.reversename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Program version
VERSION = "0.2.1"
PROGRAM_NAME = "DNS Response Monitor"


@dataclass
class DNSResult:
    """Result for a single DNS query against a server."""
    server: str
    server_ip: Optional[str]
    short_name: Optional[str]
    iteration: int
    query_time_ms: Optional[float]  # Response time in milliseconds
    status: str  # OK, TIMEOUT, REFUSED, ERROR
    error_message: Optional[str] = None


@dataclass
class DNSServerSummary:
    """Aggregated results for one DNS server across all iterations."""
    server: str
    server_ip: Optional[str]
    short_name: Optional[str]
    total_queries: int
    successful_queries: int
    failed_queries: int
    min_ms: Optional[float]
    max_ms: Optional[float]
    avg_ms: Optional[float]
    median_ms: Optional[float]
    stddev_ms: Optional[float]
    failure_rate: float  # 0.0 to 1.0
    is_additional_server: bool = False
    all_times: list = field(default_factory=list)
    errors: list = field(default_factory=list)


@dataclass
class Config:
    """Configuration data model."""
    servers_file: Optional[Path]
    additional_servers_file: Optional[Path]
    output_file: Optional[Path]
    iterations: int
    timeout: int
    query_domain: str  # Domain to query for testing
    parallel_limit: int
    verbose: bool
    variance_threshold_ms: float  # Flag servers with avg response > this
    skip_after_failures: int = 5  # Stop querying a server after this many consecutive failures


def query_dns_server(server_ip: str, query_domain: str, timeout: int = 5) -> tuple[Optional[float], str, Optional[str]]:
    """
    Send a DNS query to a specific server and measure response time.

    Args:
        server_ip: IP address of DNS server to query
        query_domain: Domain name to look up
        timeout: Query timeout in seconds

    Returns:
        Tuple of (response_time_ms, status, error_message)
    """
    try:
        # Build DNS query packet
        txn_id = random.randint(0, 65535)
        header = struct.pack('!HHHHHH', txn_id, 0x0100, 1, 0, 0, 0)

        # Encode domain name
        question = b''
        for label in query_domain.split('.'):
            question += bytes([len(label)]) + label.encode('ascii')
        question += b'\x00'  # Root label
        question += struct.pack('!HH', 1, 1)  # Type A, Class IN

        packet = header + question

        # Send via UDP
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.settimeout(timeout)

        start = time.perf_counter()
        sock.sendto(packet, (server_ip, 53))
        resp, addr = sock.recvfrom(4096)
        elapsed_ms = (time.perf_counter() - start) * 1000.0

        sock.close()

        # Validate response
        if len(resp) < 12:
            return None, 'ERROR', 'Response too short'

        resp_id = struct.unpack('!H', resp[:2])[0]
        resp_flags = struct.unpack('!H', resp[2:4])[0]

        if resp_id != txn_id:
            return None, 'ERROR', 'Transaction ID mismatch'

        if not (resp_flags & 0x8000):
            return None, 'ERROR', 'Not a response'

        # Check RCODE (bits 0-3 of flags)
        rcode = resp_flags & 0x000f
        if rcode == 5:
            return None, 'REFUSED', 'Query refused'
        elif rcode == 2:
            return None, 'ERROR', 'Server failure (SERVFAIL)'

        # TC flag (bit 9) means truncated — server responded but answer was too large for UDP
        # Still counts as OK since the server is responding
        return round(elapsed_ms, 3), 'OK', None

    except socket.timeout:
        return None, 'TIMEOUT', f'Timeout after {timeout}s'
    except ConnectionRefusedError:
        return None, 'REFUSED', 'Connection refused'
    except OSError as e:
        return None, 'ERROR', str(e)
    except Exception as e:
        return None, 'ERROR', str(e)


def get_all_domain_ips(domain: str, timeout: int = 10) -> List[str]:
    """Get all A records for a domain. Falls back to TCP for large responses."""
    logger = logging.getLogger(__name__)
    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = timeout
        resolver.lifetime = timeout

        # Try UDP first, fall back to TCP for large record sets
        try:
            answers = resolver.resolve(domain, 'A')
        except (dns.resolver.NoAnswer, dns.exception.Timeout):
            logger.debug(f"UDP failed for {domain}, trying TCP")
            answers = resolver.resolve(domain, 'A', tcp=True)

        ips = [str(rdata) for rdata in answers]
        logger.info(f"Found {len(ips)} A records for {domain}")
        return ips
    except Exception as e:
        # Final fallback: try TCP directly
        try:
            resolver2 = dns.resolver.Resolver()
            resolver2.timeout = timeout
            resolver2.lifetime = timeout
            answers = resolver2.resolve(domain, 'A', tcp=True)
            ips = [str(rdata) for rdata in answers]
            logger.info(f"Found {len(ips)} A records for {domain} (via TCP)")
            return ips
        except Exception as e2:
            logger.error(f"DNS lookup failed for {domain}: {e2}")
            return []


def perform_reverse_dns(ip: str, domain_suffix: str = "tgna.tegna.com", timeout: int = 5) -> tuple[Optional[str], Optional[str]]:
    """Reverse DNS lookup. Returns (full_hostname, short_name)."""
    try:
        resolver = dns.resolver.Resolver()
        resolver.timeout = timeout
        resolver.lifetime = timeout
        reverse_name = dns.reversename.from_address(ip)
        answers = resolver.resolve(reverse_name, 'PTR')
        if answers:
            full = str(answers[0]).rstrip('.')
            short = full
            suffix = f'.{domain_suffix}'
            if full.endswith(suffix):
                short = full[:-len(suffix)]
            return full, short
    except Exception:
        pass
    return None, None


def parse_additional_servers_csv(file_path: Path) -> List[tuple[str, str]]:
    """Parse CSV with server and optional short_name columns. Returns list of (server, short_name)."""
    logger = logging.getLogger(__name__)
    servers = []
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames or 'server' not in reader.fieldnames:
                logger.error(f"CSV missing 'server' column")
                return []
            has_name = 'short_name' in reader.fieldnames
            for row in reader:
                server = row.get('server', '').strip()
                if not server:
                    continue
                short = row.get('short_name', '').strip() if has_name else ''
                if not short:
                    short = server
                servers.append((server, short))
        logger.info(f"Parsed {len(servers)} additional servers from {file_path}")
    except Exception as e:
        logger.error(f"Error reading {file_path}: {e}")
    return servers


def run_iterations(servers: list, config: Config, ini_config: dict) -> dict:
    """
    Run DNS queries against all servers for the configured number of iterations.
    Each iteration queries all active servers in parallel using a thread pool.
    Servers that fail consecutively are removed from the active rotation.

    Returns dict mapping server_ip to DNSServerSummary.
    """
    logger = logging.getLogger(__name__)
    iterations = config.iterations
    skip_threshold = config.skip_after_failures
    max_workers = min(config.parallel_limit, len(servers))

    # Build server list with metadata
    server_entries = []
    for item in servers:
        if isinstance(item, tuple):
            ip, short, is_additional = item
        else:
            ip, short, is_additional = item, item, False
        server_entries.append((ip, short, is_additional))

    # Initialize results tracking
    results = {}
    consecutive_failures = {}
    skipped_servers = set()

    for ip, short, is_additional in server_entries:
        results[ip] = DNSServerSummary(
            server=ip, server_ip=ip, short_name=short,
            total_queries=0, successful_queries=0, failed_queries=0,
            min_ms=None, max_ms=None, avg_ms=None, median_ms=None,
            stddev_ms=None, failure_rate=0.0, is_additional_server=is_additional,
            all_times=[], errors=[]
        )
        consecutive_failures[ip] = 0

    logger.info(f"Starting {iterations} iterations across {len(server_entries)} servers "
                f"({max_workers} parallel workers, skip after {skip_threshold} failures)")
    completed = 0
    skipped_count = 0

    for iteration in range(1, iterations + 1):
        active_entries = [(ip, short, ia) for ip, short, ia in server_entries if ip not in skipped_servers]
        if iteration % 10 == 0 or iteration == 1:
            logger.info(f"Iteration {iteration}/{iterations} ({len(active_entries)} active, {len(skipped_servers)} skipped)")

        if not active_entries:
            logger.warning(f"All servers skipped after {iteration-1} iterations")
            break

        # Query all active servers in parallel
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {}
            for ip, short, is_additional in active_entries:
                future = executor.submit(query_dns_server, ip, config.query_domain, config.timeout)
                futures[future] = (ip, short)

            for future in as_completed(futures):
                ip, short = futures[future]
                completed += 1
                try:
                    elapsed_ms, status, error = future.result()
                except Exception as e:
                    elapsed_ms, status, error = None, 'ERROR', str(e)

                summary = results[ip]
                summary.total_queries += 1

                if status == 'OK' and elapsed_ms is not None:
                    summary.successful_queries += 1
                    summary.all_times.append(elapsed_ms)
                    consecutive_failures[ip] = 0
                else:
                    summary.failed_queries += 1
                    summary.errors.append(f"Iter {iteration}: {status} - {error}")
                    consecutive_failures[ip] += 1

                    if skip_threshold > 0 and consecutive_failures[ip] >= skip_threshold:
                        skipped_servers.add(ip)
                        logger.warning(f"Skipping {short} ({ip}): {skip_threshold} consecutive failures")

    # Calculate statistics
    for ip, summary in results.items():
        times = summary.all_times
        if times:
            summary.min_ms = round(min(times), 3)
            summary.max_ms = round(max(times), 3)
            summary.avg_ms = round(statistics.mean(times), 3)
            summary.median_ms = round(statistics.median(times), 3)
            if len(times) > 1:
                summary.stddev_ms = round(statistics.stdev(times), 3)
            else:
                summary.stddev_ms = 0.0
        summary.failure_rate = round(summary.failed_queries / summary.total_queries, 4) if summary.total_queries > 0 else 0.0

    logger.info(f"Completed {completed} queries, {len(skipped_servers)} servers removed from rotation")
    return results


def write_xlsx_report(results: dict, output_path: Path, config: Config, ini_config: dict) -> None:
    """Write results to XLSX with formatting."""
    logger = logging.getLogger(__name__)

    try:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "DNS Response Report"

        headers = [
            'Server', 'Server IP', 'Short Name', 'Total Queries',
            'Successful', 'Failed', 'Failure Rate %',
            'Min (ms)', 'Max (ms)', 'Avg (ms)', 'Median (ms)', 'StdDev (ms)',
            'Status', 'Sample Errors'
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Sort: non-OK first (FAILED, DEGRADED, SLOW), then OK by avg response time descending
        threshold = config.variance_threshold_ms

        def _sort_key(r):
            if r.failure_rate >= 1.0:
                return (0, 0)  # FAILED first
            elif r.failure_rate > 0:
                return (1, -r.failure_rate)  # DEGRADED next, worst first
            elif r.avg_ms and r.avg_ms > threshold:
                return (2, -(r.avg_ms or 0))  # SLOW next, slowest first
            else:
                return (3, -(r.avg_ms or 0))  # OK last

        sorted_results = sorted(results.values(), key=_sort_key)

        for row_num, summary in enumerate(sorted_results, 2):
            status = 'OK'
            if summary.failure_rate >= 1.0:
                status = 'FAILED'
            elif summary.failure_rate > 0:
                status = 'DEGRADED'
            elif summary.avg_ms and summary.avg_ms > threshold:
                status = 'SLOW'

            sample_errors = '; '.join(summary.errors[:3])
            if len(summary.errors) > 3:
                sample_errors += f' (+{len(summary.errors)-3} more)'

            row_data = [
                summary.server,
                summary.server_ip or '',
                summary.short_name or '',
                summary.total_queries,
                summary.successful_queries,
                summary.failed_queries,
                round(summary.failure_rate * 100, 2),
                summary.min_ms if summary.min_ms is not None else '',
                summary.max_ms if summary.max_ms is not None else '',
                summary.avg_ms if summary.avg_ms is not None else '',
                summary.median_ms if summary.median_ms is not None else '',
                summary.stddev_ms if summary.stddev_ms is not None else '',
                status,
                sample_errors
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

            # Status formatting (column 13)
            status_cell = ws.cell(row=row_num, column=13)
            if status == 'OK':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif status == 'SLOW':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(bold=True)

            # Avg response time highlighting (column 10)
            avg_cell = ws.cell(row=row_num, column=10)
            if summary.avg_ms is not None and summary.avg_ms > threshold:
                avg_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                avg_cell.font = Font(color="FFFFFF", bold=True)

            # Failure rate highlighting (column 7)
            fail_cell = ws.cell(row=row_num, column=7)
            if summary.failure_rate > 0:
                fail_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                fail_cell.font = Font(bold=True)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            try:
                col_letter = column[0].column_letter
            except AttributeError:
                continue
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter') and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

        workbook.save(output_path)
        logger.info(f"Wrote {len(results)} server results to {output_path}")

    except Exception as e:
        logger.error(f"Failed to write XLSX: {e}")
        raise


def format_summary(results: dict, config: Config) -> str:
    """Generate summary text."""
    summaries = list(results.values())
    total_servers = len(summaries)
    healthy = sum(1 for s in summaries if s.failure_rate == 0 and (s.avg_ms is None or s.avg_ms <= config.variance_threshold_ms))
    slow = sum(1 for s in summaries if s.avg_ms and s.avg_ms > config.variance_threshold_ms and s.failure_rate < 1.0)
    degraded = sum(1 for s in summaries if 0 < s.failure_rate < 1.0)
    failed = sum(1 for s in summaries if s.failure_rate >= 1.0)

    all_avgs = [s.avg_ms for s in summaries if s.avg_ms is not None]

    lines = [
        "=" * 60,
        "DNS RESPONSE MONITOR SUMMARY",
        "=" * 60,
        f"Hostname: {socket.gethostname()}",
        f"Program: {PROGRAM_NAME} v{VERSION}",
        "",
        f"Servers tested: {total_servers}",
        f"Iterations per server: {config.iterations}",
        f"Query domain: {config.query_domain}",
        "",
        f"Healthy: {healthy}",
        f"Slow (>{config.variance_threshold_ms}ms avg): {slow}",
        f"Degraded (partial failures): {degraded}",
        f"Failed (100% failure): {failed}",
    ]

    if all_avgs:
        lines.extend([
            "",
            "Response time statistics (across all servers):",
            f"  Fastest avg: {min(all_avgs):.3f}ms",
            f"  Slowest avg: {max(all_avgs):.3f}ms",
            f"  Overall avg: {statistics.mean(all_avgs):.3f}ms",
        ])

    # List problem servers
    problems = [s for s in summaries if s.failure_rate > 0 or (s.avg_ms and s.avg_ms > config.variance_threshold_ms)]
    if problems:
        lines.extend(["", "Servers requiring attention:"])
        for s in sorted(problems, key=lambda x: x.failure_rate, reverse=True):
            if s.failure_rate >= 1.0:
                lines.append(f"  {s.short_name} ({s.server}): FAILED - 100% failure rate")
            elif s.failure_rate > 0:
                lines.append(f"  {s.short_name} ({s.server}): DEGRADED - {s.failure_rate*100:.1f}% failures, avg {s.avg_ms:.1f}ms")
            else:
                lines.append(f"  {s.short_name} ({s.server}): SLOW - avg {s.avg_ms:.1f}ms")

    lines.append("=" * 60)
    return "\n".join(lines)


def send_email_notification(summary_text: str, xlsx_path: Path, results: dict,
                            config: Config, ini_config: dict) -> None:
    """Send email notification."""
    logger = logging.getLogger(__name__)
    if not ini_config.get('send_email', False):
        return

    try:
        summaries = list(results.values())
        failed = sum(1 for s in summaries if s.failure_rate >= 1.0)
        degraded = sum(1 for s in summaries if 0 < s.failure_rate < 1.0)
        slow = sum(1 for s in summaries if s.avg_ms and s.avg_ms > config.variance_threshold_ms and s.failure_rate == 0)

        has_error = failed > 0 or degraded > 0
        prefix = "DNS ERROR" if has_error else "DNS REPORT"

        issues = []
        if failed > 0:
            issues.append(f"{failed} failed")
        if degraded > 0:
            issues.append(f"{degraded} degraded")
        if slow > 0:
            issues.append(f"{slow} slow")

        status_part = ', '.join(issues) if issues else "all DNS servers healthy"
        domain = ini_config.get('default_discovery_domain', 'unknown')
        subject = f"{prefix} {domain} - {status_part}"

        msg = MIMEMultipart()
        msg['From'] = ini_config.get('from_email', 'dns-monitor@tgna.tegna.com')
        msg['To'] = ini_config.get('to_email', 'moldham@tegna.com')
        msg['Subject'] = subject

        if has_error:
            msg['X-Priority'] = '1'
            msg['X-MSMail-Priority'] = 'High'
            msg['Importance'] = 'High'

        msg.attach(MIMEText(summary_text, 'plain'))

        if xlsx_path.exists():
            with open(xlsx_path, 'rb') as f:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename= {xlsx_path.name}')
                msg.attach(part)

        smtp_server = ini_config.get('smtp_server', 'relay.tgna.tegna.com')
        smtp_port = ini_config.get('smtp_port', 25)

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            if ini_config.get('smtp_use_tls', False):
                server.starttls()
            username = ini_config.get('smtp_username', '')
            password = ini_config.get('smtp_password', '')
            if username and password:
                server.login(username, password)
            server.send_message(msg)
            logger.info(f"Email sent: {subject}")

    except Exception as e:
        logger.error(f"Failed to send email: {e}")


def load_configuration(config_file: str = "dns_monitor.ini") -> dict:
    """Load configuration from INI file with defaults."""
    logger = logging.getLogger(__name__)
    defaults = {
        'default_discovery_domain': 'tgna.tegna.com',
        'fallback_servers': ['8.8.8.8', '8.8.4.4', '1.1.1.1'],
        'default_parallel_limit': 10,
        'default_timeout': 2,
        'default_iterations': 100,
        'query_domain': 'ntp1.tgna.tegna.com',
        'variance_threshold_ms': 50,
        'skip_after_failures': 5,
        'output_directory': '.\\Reports',
        'send_email': False,
        'smtp_server': '',
        'smtp_port': 25,
        'smtp_use_tls': False,
        'smtp_username': '',
        'smtp_password': '',
        'from_email': '',
        'to_email': '',
    }

    config = configparser.ConfigParser()
    try:
        if Path(config_file).exists():
            config.read(config_file)
            logger.debug(f"Loaded config from {config_file}")

            if 'dns_settings' in config:
                s = config['dns_settings']
                defaults['default_discovery_domain'] = s.get('default_discovery_domain', defaults['default_discovery_domain'])
                fallback_str = s.get('fallback_servers', '')
                if fallback_str:
                    defaults['fallback_servers'] = [x.strip() for x in fallback_str.split(',')]
                defaults['query_domain'] = s.get('query_domain', defaults['query_domain'])

            if 'monitor_settings' in config:
                s = config['monitor_settings']
                defaults['default_parallel_limit'] = s.getint('default_parallel_limit', defaults['default_parallel_limit'])
                defaults['default_timeout'] = s.getint('default_timeout', defaults['default_timeout'])
                defaults['default_iterations'] = s.getint('default_iterations', defaults['default_iterations'])
                defaults['variance_threshold_ms'] = s.getfloat('variance_threshold_ms', defaults['variance_threshold_ms'])
                defaults['skip_after_failures'] = s.getint('skip_after_failures', defaults['skip_after_failures'])
                defaults['output_directory'] = s.get('output_directory', defaults['output_directory'])

            if 'email_settings' in config:
                s = config['email_settings']
                defaults['send_email'] = s.getboolean('send_email', defaults['send_email'])
                defaults['smtp_server'] = s.get('smtp_server', defaults['smtp_server'])
                defaults['smtp_port'] = s.getint('smtp_port', defaults['smtp_port'])
                defaults['smtp_use_tls'] = s.getboolean('smtp_use_tls', defaults['smtp_use_tls'])
                defaults['smtp_username'] = s.get('smtp_username', defaults['smtp_username'])
                defaults['smtp_password'] = s.get('smtp_password', defaults['smtp_password'])
                defaults['from_email'] = s.get('from_email', defaults['from_email'])
                defaults['to_email'] = s.get('to_email', defaults['to_email'])
        else:
            logger.debug(f"Config file {config_file} not found, using defaults")
    except Exception as e:
        logger.warning(f"Error reading config: {e}")

    return defaults


def setup_logging(verbose: bool = False) -> None:
    """Configure logging."""
    log_level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(level=log_level, format='%(asctime)s - %(levelname)s - %(message)s',
                        handlers=[logging.StreamHandler(sys.stdout)])


def parse_arguments() -> Config:
    """Parse CLI arguments."""
    # First pass: check for --config to load the right INI file
    pre_parser = argparse.ArgumentParser(add_help=False)
    pre_parser.add_argument('-c', '--config', type=str, default='dns_monitor.ini',
                            help='Path to INI configuration file')
    pre_args, _ = pre_parser.parse_known_args()

    ini_config = load_configuration(pre_args.config)

    parser = argparse.ArgumentParser(
        description='DNS Response Monitor - Test DNS server response times',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Examples:
  Default (discover DCs from {ini_config['default_discovery_domain']}):
    %(prog)s

  With a different config file:
    %(prog)s -c my_dns_monitor.ini

  With additional servers CSV:
    %(prog)s --additional-servers dns_servers.csv

  Custom iterations and timeout:
    %(prog)s -n 50 -t 3

  With server list file:
    %(prog)s -s servers.txt -n 200
        """)

    parser.add_argument('-c', '--config', type=str, default='dns_monitor.ini', metavar='FILE',
                        help=f'Path to INI configuration file [default: dns_monitor.ini]')
    parser.add_argument('-s', '--servers-file', type=Path, metavar='FILE',
                        help=f'Server list file (default: discover from {ini_config["default_discovery_domain"]})')
    parser.add_argument('--additional-servers', type=Path, metavar='FILE',
                        help='CSV with additional DNS servers. Auto-detects dns_servers.csv if present.')
    parser.add_argument('-n', '--iterations', type=int,
                        default=ini_config['default_iterations'], metavar='N',
                        help=f'Number of test iterations per server [default: {ini_config["default_iterations"]}]')
    parser.add_argument('-t', '--timeout', type=int,
                        default=ini_config['default_timeout'], metavar='SECONDS',
                        help=f'DNS query timeout [default: {ini_config["default_timeout"]}]')
    parser.add_argument('-q', '--query-domain', type=str,
                        default=ini_config['query_domain'], metavar='DOMAIN',
                        help=f'Domain to query for testing [default: {ini_config["query_domain"]}]')
    parser.add_argument('-p', '--parallel-limit', type=int,
                        default=ini_config['default_parallel_limit'], metavar='N',
                        help=f'Max concurrent workers [default: {ini_config["default_parallel_limit"]}]')
    parser.add_argument('-o', '--output-file', type=Path, metavar='FILE',
                        help='Output XLSX file path (default: auto-generated)')
    parser.add_argument('--skip-after', type=int,
                        default=ini_config['skip_after_failures'], metavar='N',
                        help=f'Stop querying a server after N consecutive failures (0 to disable) [default: {ini_config["skip_after_failures"]}]')
    parser.add_argument('-v', '--verbose', action='store_true',
                        help='Enable verbose logging')
    parser.add_argument('--version', action='version',
                        version=f'{PROGRAM_NAME} {VERSION}')

    args = parser.parse_args()

    # Auto-detect additional servers CSV
    additional = args.additional_servers
    if additional is None:
        default_csv = Path('dns_servers.csv')
        if default_csv.exists():
            additional = default_csv

    return Config(
        servers_file=args.servers_file,
        additional_servers_file=additional,
        output_file=args.output_file,
        iterations=args.iterations,
        timeout=args.timeout,
        query_domain=args.query_domain,
        parallel_limit=args.parallel_limit,
        verbose=args.verbose,
        variance_threshold_ms=ini_config['variance_threshold_ms'],
        skip_after_failures=args.skip_after
    )


def main():
    """Main entry point."""
    try:
        ini_config = load_configuration()
        config = parse_arguments()
        setup_logging(config.verbose)

        logger = logging.getLogger(__name__)
        logger.info(f"{PROGRAM_NAME} v{VERSION} starting...")
        logger.info(f"Iterations: {config.iterations}, Timeout: {config.timeout}s")
        logger.info(f"Query domain: {config.query_domain}")

        # Discover or load primary DNS servers
        all_servers = []  # List of (ip, short_name, is_additional)

        if config.servers_file:
            logger.info(f"Loading servers from {config.servers_file}")
            with open(config.servers_file, 'r') as f:
                for line in f:
                    server = line.strip()
                    if server:
                        all_servers.append((server, server, False))
        else:
            domain = ini_config['default_discovery_domain']
            logger.info(f"Discovering DNS servers from {domain}")
            ips = get_all_domain_ips(domain, config.timeout)
            if not ips:
                logger.warning(f"No A records for {domain}, using fallback servers")
                ips = ini_config['fallback_servers']

            # Reverse DNS for short names (parallel)
            logger.info(f"Resolving short names for {len(ips)} servers...")
            with ThreadPoolExecutor(max_workers=min(20, len(ips))) as executor:
                future_to_ip = {executor.submit(perform_reverse_dns, ip, domain, config.timeout): ip for ip in ips}
                for future in as_completed(future_to_ip):
                    ip = future_to_ip[future]
                    try:
                        full, short = future.result()
                    except Exception:
                        full, short = None, None
                    all_servers.append((ip, short or ip, False))

            logger.info(f"Discovered {len(all_servers)} primary DNS servers")

        # Load additional servers
        if config.additional_servers_file:
            logger.info(f"Loading additional servers from {config.additional_servers_file}")
            additional = parse_additional_servers_csv(config.additional_servers_file)
            for server, short in additional:
                # Resolve hostname to IP if needed
                try:
                    ip = socket.gethostbyname(server)
                except socket.gaierror:
                    ip = server
                all_servers.append((ip, short, True))

        if not all_servers:
            logger.error("No DNS servers to test")
            return 1

        logger.info(f"Testing {len(all_servers)} total servers with {config.iterations} iterations each")

        # Run iterations
        results = run_iterations(all_servers, config, ini_config)

        # Determine output path
        if config.output_file:
            output_path = config.output_file
        else:
            output_dir = Path(ini_config.get('output_directory', '.\\Reports'))
            output_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            output_path = output_dir / f"dns_monitor_report_{timestamp}.xlsx"

        # Write report
        write_xlsx_report(results, output_path, config, ini_config)

        # Summary
        summary = format_summary(results, config)
        print(summary)

        summary_path = output_path.with_suffix('.txt')
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(summary + '\n')
        logger.info(f"Summary written to {summary_path}")

        # Email
        send_email_notification(summary, output_path, results, config, ini_config)

        # Exit code
        has_failures = any(s.failure_rate > 0 for s in results.values())
        if has_failures:
            logger.warning("Completed with issues")
            return 1

        logger.info("Completed successfully")
        return 0

    except KeyboardInterrupt:
        print("\nCancelled by user", file=sys.stderr)
        return 1
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
