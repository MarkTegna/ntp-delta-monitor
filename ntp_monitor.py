#!/usr/bin/env python3
"""
NTP Delta Monitor - A Windows-based NTP monitoring program
Queries multiple NTP sources for time synchronization analysis
"""

import argparse
import configparser
import csv
import logging
import os
import smtplib
import socket
import ssl
import statistics
import struct
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from enum import Enum
from pathlib import Path
from typing import Optional, List

import dns.resolver
import dns.reversename
import ntplib
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Program version
VERSION = "3.5.0"
PROGRAM_NAME = "NTP Delta Monitor"


class NTPStatus(Enum):
    """
    Status enumeration for NTP query results.

    Implements error classification system as required by:
    - Requirements 2.5, 8.1, 8.2, 8.3: Different types of NTP failures

    Values:
    - OK: Successful NTP query with synchronized server
    - UNREACHABLE: Server cannot be reached (DNS failure, connection refused)
    - TIMEOUT: Query exceeded specified timeout period
    - UNSYNCHRONIZED: Server responded but is not synchronized (stratum >= 16)
    - ERROR: NTP protocol errors or other unexpected failures
    """
    OK = "OK"
    UNREACHABLE = "UNREACHABLE"
    TIMEOUT = "TIMEOUT"
    UNSYNCHRONIZED = "UNSYNCHRONIZED"
    ERROR = "ERROR"


@dataclass
class NTPResponse:
    """Data model for NTP server response"""
    timestamp_utc: datetime
    query_rtt_ms: float
    stratum: int
    root_delay_ms: float
    root_dispersion_ms: float
    is_synchronized: bool
    offset_seconds: float  # NTP offset (time difference accounting for network delay)
    leap_indicator: int = 0  # Leap second warning: 0=none, 1=+1s, 2=-1s, 3=unsync
    precision: int = 0  # Clock precision as power of 2 (e.g., -20 = ~1μs)
    ref_id: str = ''  # Reference ID (upstream source - GPS/PPS for stratum 1, IP for stratum 2+)
    ref_time: Optional[datetime] = None  # When server clock was last set/corrected
    poll_interval: int = 0  # Poll interval as power of 2 seconds
    error_message: Optional[str] = None


def format_delta_value(delta_seconds: Optional[float], format_type: str) -> Optional[float]:
    """
    Convert delta value to specified format with appropriate precision.

    Args:
        delta_seconds: Delta value in seconds (None if calculation failed)
        format_type: 'seconds' for decimal seconds, 'milliseconds' for integer milliseconds

    Returns:
        Formatted delta value according to format_type, None if input is None
    """
    if delta_seconds is None:
        return None

    if format_type == 'seconds':
        # Return decimal seconds with millisecond precision (3 decimal places)
        return round(delta_seconds, 3)
    elif format_type == 'milliseconds':
        # Convert to integer milliseconds
        return int(round(delta_seconds * 1000))
    else:
        # Default to seconds format if unknown format specified
        return round(delta_seconds, 3)


@dataclass
class NTPResult:
    """Data model for complete NTP query result"""
    timestamp_utc: datetime
    ntp_server: str
    ntp_server_ip: Optional[str]
    hostname: Optional[str]  # Full hostname from reverse DNS
    short_name: Optional[str]  # Short name (hostname without tgna.tegna.com suffix)
    ntp_time_utc: Optional[datetime]
    query_rtt_ms: Optional[float]
    stratum: Optional[int]
    root_delay_ms: Optional[float]
    root_dispersion_ms: Optional[float]
    delta_seconds: Optional[float]  # vs reference NTP source (raw seconds)
    delta_formatted: Optional[float]  # formatted according to config (seconds/milliseconds)
    status: NTPStatus
    error_message: Optional[str] = None
    is_additional_server: bool = False  # True if from additional servers CSV (excluded from error counts)
    failure_count: int = 0  # Consecutive failure count for this server
    missed: int = 0  # Total connection error count (never resets)
    failed: str = 'ok'  # Status: 'failed' or 'ok'
    time: int = 0  # Consecutive variance threshold violations (resets when within threshold)
    leap_indicator: Optional[int] = None  # Leap second warning
    precision: Optional[int] = None  # Clock precision as power of 2
    ref_id: Optional[str] = None  # Reference source identifier
    ref_time: Optional[datetime] = None  # Last clock correction time
    poll_interval: Optional[int] = None  # Poll interval as power of 2 seconds
    kerberos_status: Optional[str] = None  # Kerberos KDC check result
    dns_status: Optional[str] = None  # DNS service check result
    ldap_status: Optional[str] = None  # LDAP service check result (port 389)
    ldaps_status: Optional[str] = None  # LDAPS service check result (port 636)
    ldaps_cert_expiry: Optional[str] = None  # LDAPS certificate expiration date


@dataclass
class Config:
    """Configuration data model for NTP monitor"""
    reference_ntp: str  # Primary reference NTP source
    ntp_servers_file: Optional[Path]  # Path to NTP server list file (None for auto-discovery)
    additional_servers_file: Optional[Path]  # Path to additional servers CSV file (excluded from error counts)
    output_file: Optional[Path]
    format_type: str  # 'seconds' or 'milliseconds'
    parallel_limit: int
    ntp_timeout: int
    verbose: bool
    skip_threshold: int = 10  # Skip additional servers with failure_count >= this value


@dataclass
class ExtendedServerRecord:
    """
    Extended server record with all tracking columns.

    Implements data structure requirements from:
    - Requirement 7.1: Contains exactly 7 fields
    - Requirement 7.2: server is non-empty string
    - Requirement 7.3: failure_count is non-negative integer
    - Requirement 7.4: missed is non-negative integer
    - Requirement 7.5: failed is either 'ok' or 'failed'
    - Requirement 7.6: short_name can be any string including empty
    - Requirement 7.7: dlist can be any string
    - Requirement 7.8: time is non-negative integer (consecutive variance violations)
    """
    server: str                    # NTP server hostname or IP
    short_name: str                # Display name (can be empty)
    failure_count: int             # Consecutive failures (resets on success)
    missed: int                    # Total connection errors (never resets)
    failed: str                    # Status: 'failed' or 'ok'
    dlist: str                     # Distribution list email address
    time: int                      # Consecutive variance threshold violations (resets when within threshold)

    def __post_init__(self):
        """
        Validate field values after initialization.

        Validates:
        - Requirement 7.2: server must be non-empty string
        - Requirement 7.3: failure_count must be non-negative integer
        - Requirement 7.4: missed must be non-negative integer
        - Requirement 7.5: failed must be 'ok' or 'failed'
        - Requirement 7.8: time must be non-negative integer
        """
        # Validate server is non-empty
        if not self.server or not isinstance(self.server, str):
            raise ValueError(f"server must be a non-empty string, got: {self.server!r}")

        # Validate failure_count is non-negative integer
        if not isinstance(self.failure_count, int) or self.failure_count < 0:
            raise ValueError(f"failure_count must be a non-negative integer, got: {self.failure_count!r}")

        # Validate missed is non-negative integer
        if not isinstance(self.missed, int) or self.missed < 0:
            raise ValueError(f"missed must be a non-negative integer, got: {self.missed!r}")

        # Validate failed status
        if self.failed not in ('ok', 'failed'):
            raise ValueError(f"failed must be 'ok' or 'failed', got: {self.failed!r}")

        # Validate time is non-negative integer
        if not isinstance(self.time, int) or self.time < 0:
            raise ValueError(f"time must be a non-negative integer, got: {self.time!r}")


@dataclass
class SummaryStats:
    """Data model for summary statistics"""
    total_servers: int
    successful_servers: int
    failed_servers: int
    min_delta: Optional[float]
    max_delta: Optional[float]
    avg_delta: Optional[float]
    status_counts: dict  # Count of each status type
    has_errors: bool  # True if any ERROR/TIMEOUT/UNSYNCHRONIZED status

    @dataclass
    class ExtendedServerRecord:
        """
        Extended server record with all tracking columns.

        Implements data structure requirements from:
        - Requirement 7.1: Contains exactly 6 fields
        - Requirement 7.2: server is non-empty string
        - Requirement 7.3: failure_count is non-negative integer
        - Requirement 7.4: missed is non-negative integer
        - Requirement 7.5: failed is either 'ok' or 'failed'
        - Requirement 7.6: short_name can be any string including empty
        - Requirement 7.7: dlist can be any string
        """
        server: str                    # NTP server hostname or IP
        short_name: str                # Display name (can be empty)
        failure_count: int             # Consecutive failures (resets on success)
        missed: int                    # Total connection errors (never resets)
        failed: str                    # Status: 'failed' or 'ok'
        dlist: str                     # Distribution list email address

        def __post_init__(self):
            """
            Validate field values after initialization.

            Validates:
            - Requirement 7.2: server must be non-empty string
            - Requirement 7.3: failure_count must be non-negative integer
            - Requirement 7.4: missed must be non-negative integer
            - Requirement 7.5: failed must be 'ok' or 'failed'
            """
            # Validate server is non-empty
            if not self.server or not isinstance(self.server, str):
                raise ValueError(f"server must be a non-empty string, got: {self.server!r}")

            # Validate failure_count is non-negative integer
            if not isinstance(self.failure_count, int) or self.failure_count < 0:
                raise ValueError(f"failure_count must be a non-negative integer, got: {self.failure_count!r}")

            # Validate missed is non-negative integer
            if not isinstance(self.missed, int) or self.missed < 0:
                raise ValueError(f"missed must be a non-negative integer, got: {self.missed!r}")

            # Validate failed status
            if self.failed not in ('ok', 'failed'):
                raise ValueError(f"failed must be 'ok' or 'failed', got: {self.failed!r}")


    @dataclass
    class SummaryStats:
        """Data model for summary statistics"""
        total_servers: int
        successful_servers: int
        failed_servers: int
        min_delta: Optional[float]
        max_delta: Optional[float]
        avg_delta: Optional[float]


def query_ntp_server(hostname: str, timeout: int = 30) -> NTPResponse:
    """
    Query a single NTP server and return structured response data.

    Args:
        hostname: NTP server hostname or IP address
        timeout: Query timeout in seconds

    Returns:
        NTPResponse with timestamp, RTT, stratum, delays, and dispersion

    Raises:
        Exception: For NTP protocol errors, timeouts, or network issues
    """
    logger = logging.getLogger(__name__)

    try:
        # Create NTP client and query server
        ntp_client = ntplib.NTPClient()
        logger.debug(f"Querying NTP server: {hostname} (timeout: {timeout}s)")

        # Log timing information for verbose mode
        start_time = datetime.now(timezone.utc)

        # Query using NTP version 4
        response = ntp_client.request(hostname, version=4, timeout=timeout)

        # Calculate and log timing information
        end_time = datetime.now(timezone.utc)
        query_duration = (end_time - start_time).total_seconds()
        logger.debug(f"NTP query completed in {query_duration:.3f} seconds")

        # Convert NTP timestamp to UTC datetime
        timestamp_utc = datetime.fromtimestamp(response.tx_time, tz=timezone.utc)

        # Validate timestamp format
        if not validate_timestamp_format(timestamp_utc):
            raise Exception(f"Invalid timestamp format from server: {timestamp_utc}")

        # Calculate RTT in milliseconds
        query_rtt_ms = response.delay * 1000.0

        # Convert delays and dispersion to milliseconds
        root_delay_ms = response.root_delay * 1000.0
        root_dispersion_ms = response.root_dispersion * 1000.0

        # Check if server is synchronized (stratum < 16)
        is_synchronized = response.stratum < 16

        # Log detailed NTP query information when verbose
        logger.debug(f"NTP response from {hostname}:")
        logger.debug(f"  Timestamp: {timestamp_utc.isoformat()}")
        logger.debug(f"  Stratum: {response.stratum}")
        logger.debug(f"  RTT: {query_rtt_ms:.2f}ms")
        logger.debug(f"  Root delay: {root_delay_ms:.2f}ms")
        logger.debug(f"  Root dispersion: {root_dispersion_ms:.2f}ms")
        logger.debug(f"  Synchronized: {is_synchronized}")
        logger.debug(f"  Query duration: {query_duration:.3f}s")

        # Extract additional NTP response fields
        leap_indicator = response.leap
        precision = response.precision
        poll_interval = response.poll

        # Extract reference ID - for stratum 1 it's a 4-char string, for stratum 2+ it's an IP
        ref_id = ''
        try:
            raw_ref_id = response.ref_id
            if response.stratum == 1:
                # Stratum 1: ref_id is a 4-character ASCII string (e.g., 'GPS', 'PPS', 'GOOG')
                ref_id = struct.pack('!I', int(raw_ref_id)).decode('ascii').rstrip('\x00')
            else:
                # Stratum 2+: ref_id is the IP address of the upstream server
                ref_id = socket.inet_ntoa(struct.pack('!I', int(raw_ref_id)))
        except (AttributeError, TypeError, ValueError, struct.error, OSError):
            ref_id = str(raw_ref_id) if raw_ref_id else ''

        # Extract reference timestamp (when clock was last set/corrected)
        ref_time = None
        try:
            if response.ref_time and response.ref_time > 0:
                ref_time = datetime.fromtimestamp(response.ref_time, tz=timezone.utc)
        except (ValueError, OSError, OverflowError):
            ref_time = None

        logger.debug(f"  Leap indicator: {leap_indicator}")
        logger.debug(f"  Precision: 2^{precision} seconds")
        logger.debug(f"  Reference ID: {ref_id}")
        logger.debug(f"  Reference time: {ref_time.isoformat() if ref_time else 'N/A'}")
        logger.debug(f"  Poll interval: 2^{poll_interval} seconds")

        ntp_response = NTPResponse(
            timestamp_utc=timestamp_utc,
            query_rtt_ms=query_rtt_ms,
            stratum=response.stratum,
            root_delay_ms=root_delay_ms,
            root_dispersion_ms=root_dispersion_ms,
            is_synchronized=is_synchronized,
            offset_seconds=response.offset,  # NTP offset (accounts for network delay)
            leap_indicator=leap_indicator,
            precision=precision,
            ref_id=ref_id,
            ref_time=ref_time,
            poll_interval=poll_interval,
            error_message=None
        )

        # Validate the response
        status, error_msg = validate_ntp_response(ntp_response)
        if status != NTPStatus.OK:
            ntp_response.error_message = error_msg
            logger.debug(f"NTP response validation failed: {error_msg}")
        else:
            logger.debug(f"NTP response validation successful")

        return ntp_response

    except ntplib.NTPException as e:
        logger.debug(f"NTP protocol error for {hostname}: {e}")
        raise Exception(f"NTP protocol error: {e}")

    except socket.timeout:
        logger.debug(f"Timeout querying NTP server: {hostname} (timeout: {timeout}s)")
        raise Exception(f"Timeout after {timeout} seconds")

    except socket.gaierror as e:
        logger.debug(f"DNS resolution failed for {hostname}: {e}")
        raise Exception(f"DNS resolution failed: {e}")

    except Exception as e:
        logger.debug(f"Unexpected error querying {hostname}: {e}")
        raise Exception(f"Network error: {e}")


def validate_ntp_response(response: NTPResponse) -> tuple[NTPStatus, Optional[str]]:
    """
    Validate NTP response and determine status.

    Args:
        response: NTPResponse object to validate

    Returns:
        Tuple of (NTPStatus, error_message)
    """
    # Check for stratum 16+ (unsynchronized) condition
    if response.stratum >= 16:
        return NTPStatus.UNSYNCHRONIZED, f"Server unsynchronized (stratum {response.stratum})"

    # Check if response indicates synchronization
    if not response.is_synchronized:
        return NTPStatus.UNSYNCHRONIZED, "Server reports unsynchronized state"

    # Response is valid and synchronized
    return NTPStatus.OK, None


def validate_timestamp_format(timestamp: datetime) -> bool:
    """
    Validate that timestamp is properly formatted and in UTC.

    Args:
        timestamp: datetime object to validate

    Returns:
        True if timestamp is valid UTC datetime, False otherwise
    """
    try:
        # Check if timestamp has timezone info and is UTC
        if timestamp.tzinfo is None:
            return False

        if timestamp.tzinfo != timezone.utc:
            return False

        # Check if timestamp is reasonable (not too far in past/future)
        now = datetime.now(timezone.utc)
        time_diff = abs((timestamp - now).total_seconds())

        # Allow up to 24 hours difference (generous for NTP sync issues)
        if time_diff > 86400:  # 24 hours in seconds
            return False

        return True

    except (AttributeError, TypeError, ValueError):
        return False


def handle_ntp_query_error(hostname: str, error: Exception, timeout: int) -> tuple[NTPStatus, str]:
    """
    Classify NTP query errors and return appropriate status and message.

    This function implements comprehensive error classification as required by:
    - Requirements 2.5: Record appropriate error status and continue with remaining servers
    - Requirements 8.1: Record TIMEOUT status for network timeouts
    - Requirements 8.2: Record UNREACHABLE status for unreachable servers
    - Requirements 8.3: Record ERROR status with specific error details

    Args:
        hostname: NTP server hostname that failed
        error: Exception that occurred during query
        timeout: Timeout value used for the query

    Returns:
        Tuple of (NTPStatus, error_message)
    """
    logger = logging.getLogger(__name__)
    error_str = str(error).lower()

    # Log error details and recovery actions when verbose
    logger.debug(f"Classifying error for {hostname}: {error}")

    # Classify different types of NTP failures with specific error messages
    if "timeout" in error_str:
        status = NTPStatus.TIMEOUT
        message = f"Query timeout after {timeout} seconds"
        logger.debug(f"Recovery action: Continuing with remaining servers after timeout")
        return status, message

    elif "dns resolution failed" in error_str or "name resolution" in error_str or "nodename nor servname provided" in error_str:
        status = NTPStatus.UNREACHABLE
        message = f"DNS resolution failed for {hostname}"
        logger.debug(f"Recovery action: Continuing with remaining servers after DNS failure")
        return status, message

    elif "connection refused" in error_str or "unreachable" in error_str or "no route to host" in error_str:
        status = NTPStatus.UNREACHABLE
        message = f"Server unreachable: {hostname}"
        logger.debug(f"Recovery action: Continuing with remaining servers after connection failure")
        return status, message

    elif "ntp protocol error" in error_str or "invalid ntp response" in error_str:
        status = NTPStatus.ERROR
        message = f"NTP protocol error: {error}"
        logger.debug(f"Recovery action: Continuing with remaining servers after protocol error")
        return status, message

    elif "network error" in error_str or "socket error" in error_str:
        status = NTPStatus.ERROR
        message = f"Network communication error: {error}"
        logger.debug(f"Recovery action: Continuing with remaining servers after network error")
        return status, message

    elif "invalid timestamp" in error_str or "malformed" in error_str:
        status = NTPStatus.ERROR
        message = f"Invalid NTP response format: {error}"
        logger.debug(f"Recovery action: Continuing with remaining servers after timestamp validation error")
        return status, message

    else:
        # Catch-all for any other unexpected errors
        status = NTPStatus.ERROR
        message = f"Unexpected error: {error}"
        logger.debug(f"Recovery action: Continuing with remaining servers after unexpected error")
        return status, message


def update_failure_count(server: str, current_count: int, status: NTPStatus) -> int:
    """
    Calculate updated failure count based on query status.

    Implements failure tracking logic as required by:
    - Requirements 3.1: Increment failure count for all failure statuses
    - Requirements 3.2: Increment for ERROR status
    - Requirements 3.3: Increment for TIMEOUT status
    - Requirements 3.4: Increment for UNREACHABLE status
    - Requirements 3.5: Increment for UNSYNCHRONIZED status
    - Requirements 4.1: Reset count to 0 for OK status
    - Requirements 4.2: Reset regardless of previous value

    Args:
        server: Server hostname for logging
        current_count: Current consecutive failure count
        status: NTP query result status

    Returns:
        Updated failure count (0 for success, current_count + 1 for failure)
    """
    logger = logging.getLogger(__name__)

    if status == NTPStatus.OK:
        # Success: reset count to 0
        new_count = 0
        logger.debug(f"Server {server}: failure count {current_count} -> {new_count} (status: {status.value})")
    elif status in [NTPStatus.ERROR, NTPStatus.TIMEOUT, NTPStatus.UNREACHABLE, NTPStatus.UNSYNCHRONIZED]:
        # Failure: increment count
        new_count = current_count + 1
        logger.debug(f"Server {server}: failure count {current_count} -> {new_count} (status: {status.value})")
    else:
        # Defensive: unknown status treated as failure
        new_count = current_count + 1
        logger.warning(f"Unknown status {status} for server {server}, treating as failure")
        logger.debug(f"Server {server}: failure count {current_count} -> {new_count} (status: {status.value})")

    return new_count


def update_tracking_counters(server: str, current_missed: int, current_failed: str,
                            status: NTPStatus, delta_seconds: Optional[float] = None,
                            variance_threshold_ms: float = 33.0, current_time: int = 0) -> tuple[int, str, int]:
    """
    Calculate updated tracking counters based on query status and variance threshold.

    This function implements the core tracking logic for the new columns:
    - Resets missed counter on successful connection (status == OK)
    - Increments missed counter on any connection error
    - Sets failed status to 'failed' when missed >= 10, 'ok' otherwise
    - Increments time counter when variance threshold is exceeded
    - Resets time counter when variance is within threshold

    Implements requirements:
    - Requirement 3.1: Increment missed counter on non-OK status
    - Requirement 3.2: Reset missed counter on OK status
    - Requirement 3.3: Log missed counter increments
    - Requirement 3.4: Log when missed counter is reset
    - Requirement 4.1: Set failed='failed' when missed >= 10
    - Requirement 4.2: Set failed='ok' when missed < 10
    - Requirement 4.3: Log failed status updates
    - Requirement 4.4: Reset failed='ok' when connection is successful
    - Requirement 5.1: Increment time counter when variance exceeds threshold
    - Requirement 5.2: Reset time counter when variance is within threshold

    Args:
        server: Server hostname for logging
        current_missed: Current total missed count
        current_failed: Current failed status ('ok' or 'failed')
        status: NTP query result status
        delta_seconds: Time delta in seconds (used for variance checking)
        variance_threshold_ms: Variance threshold in milliseconds
        current_time: Current consecutive variance violation count

    Returns:
        Tuple of (new_missed, new_failed, new_time)
        - new_missed: Reset to 0 if OK, incremented if error
        - new_failed: 'failed' if new_missed >= 10, 'ok' otherwise
        - new_time: Incremented if variance exceeded, reset to 0 if within threshold

    Preconditions:
        - server is non-empty string
        - current_missed is non-negative integer
        - current_failed is either 'ok' or 'failed'
        - status is valid NTPStatus enum value
        - current_time is non-negative integer

    Postconditions:
        - new_missed = 0 if status == OK
        - new_missed = current_missed + 1 if status != OK
        - new_failed = 'failed' if new_missed >= 10, else 'ok'
        - new_time = current_time + 1 if variance exceeded
        - new_time = 0 if variance within threshold
    """
    logger = logging.getLogger(__name__)

    # Step 1: Update missed counter based on connection status
    if status == NTPStatus.OK:
        # Reset missed counter on successful connection
        new_missed = 0
        if current_missed > 0:
            logger.info(f"Server {server}: missed count RESET to 0 (was {current_missed}) - successful connection")
        else:
            logger.debug(f"Server {server}: missed count remains 0 - successful connection")
    else:
        # Increment missed counter on error
        new_missed = current_missed + 1
        logger.info(f"Server {server}: missed count incremented to {new_missed} (connection error: {status.value})")

    # Step 2: Update failed status based on missed threshold
    if new_missed >= 10:
        new_failed = 'failed'
        if current_failed != new_failed:
            logger.info(f"Server {server}: status set to 'failed' (missed={new_missed})")
        else:
            logger.debug(f"Server {server}: status remains 'failed' (missed={new_missed})")
    else:
        new_failed = 'ok'
        if current_failed != new_failed:
            logger.info(f"Server {server}: status set to 'ok' (missed={new_missed})")
        else:
            logger.debug(f"Server {server}: status remains 'ok' (missed={new_missed})")

    # Step 3: Update time counter based on variance threshold
    new_time = current_time
    if status == NTPStatus.OK and delta_seconds is not None:
        # Only check variance for successful connections
        delta_ms = abs(delta_seconds * 1000.0)
        if delta_ms > variance_threshold_ms:
            # Variance exceeded threshold - increment time counter
            new_time = current_time + 1
            logger.info(f"Server {server}: time count incremented to {new_time} (variance {delta_ms:.2f}ms > {variance_threshold_ms}ms)")
        else:
            # Variance within threshold - reset time counter
            new_time = 0
            if current_time > 0:
                logger.info(f"Server {server}: time count RESET to 0 (was {current_time}) - variance {delta_ms:.2f}ms <= {variance_threshold_ms}ms")
            else:
                logger.debug(f"Server {server}: time count remains 0 - variance {delta_ms:.2f}ms <= {variance_threshold_ms}ms")

    return new_missed, new_failed, new_time


def check_kerberos_kdc(host: str, timeout: int = 5) -> str:
    """
    Check if a Kerberos KDC is responding by verifying port 88 (Kerberos) accepts
    connections AND port 389 (LDAP) responds to an anonymous bind request.

    The LDAP anonymous bind proves the DC is alive and processing requests — not just
    that the port is open. A valid LDAP BindResponse confirms the directory service
    is functional, which is a strong indicator the KDC is also operational since both
    services run on the same Domain Controller.

    Args:
        host: Hostname or IP address to check
        timeout: Connection timeout in seconds

    Returns:
        Status string:
        - 'OK' if Kerberos port open AND LDAP responds with valid data
        - 'KRB_ONLY' if Kerberos port open but LDAP not responding
        - 'NO_RESPONSE' if Kerberos port is closed/unreachable
        - 'ERROR: <details>' for unexpected errors
    """
    logger = logging.getLogger(__name__)

    # Step 1: Check if Kerberos port 88 accepts TCP connections
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        result = sock.connect_ex((host, 88))
        sock.close()

        if result != 0:
            logger.debug(f"Kerberos KDC {host}: port 88 closed (connect_ex={result})")
            return 'NO_RESPONSE'

        logger.debug(f"Kerberos KDC {host}: port 88 open")
    except socket.timeout:
        logger.debug(f"Kerberos KDC {host}: port 88 connection timed out")
        return 'NO_RESPONSE'
    except Exception as e:
        logger.debug(f"Kerberos KDC {host}: port 88 error: {e}")
        return f'ERROR: {e}'

    # Step 2: Verify LDAP responds to anonymous bind (proves DC is processing requests)
    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        sock.connect((host, 389))

        # Anonymous LDAP Simple Bind Request (RFC 4511)
        # MessageID=1, BindRequest(version=3, name="", simple auth="")
        bind_request = bytes([
            0x30, 0x0c,              # SEQUENCE (12 bytes)
            0x02, 0x01, 0x01,        # INTEGER MessageID = 1
            0x60, 0x07,              # APPLICATION 0 (BindRequest) (7 bytes)
            0x02, 0x01, 0x03,        # INTEGER version = 3
            0x04, 0x00,              # OCTET STRING name = "" (anonymous)
            0x80, 0x00               # CONTEXT 0 (simple auth) = "" (no password)
        ])

        sock.sendall(bind_request)
        resp = sock.recv(4096)
        sock.close()

        if resp and len(resp) > 5:
            logger.debug(f"Kerberos KDC {host}: LDAP responded with {len(resp)} bytes (DC confirmed)")
            return 'OK'

        logger.debug(f"Kerberos KDC {host}: LDAP empty response")
        return 'KRB_ONLY'

    except socket.timeout:
        logger.debug(f"Kerberos KDC {host}: LDAP timed out (port 88 open but LDAP unresponsive)")
        return 'KRB_ONLY'
    except ConnectionRefusedError:
        logger.debug(f"Kerberos KDC {host}: LDAP port 389 refused (port 88 open but no LDAP)")
        return 'KRB_ONLY'
    except OSError as e:
        logger.debug(f"Kerberos KDC {host}: LDAP error: {e}")
        return 'KRB_ONLY'
    except Exception as e:
        logger.debug(f"Kerberos KDC {host}: unexpected LDAP error: {e}")
        return f'ERROR: {e}'


def check_dns_service(host: str, timeout: int = 5) -> str:
    """
    Check if a DNS server is responding by sending a real DNS query and verifying
    the response contains valid DNS data.

    Sends an A record query for a well-known domain (the server's own domain or
    a root hint) and checks that the response is a valid DNS packet with the
    correct transaction ID. This proves the DNS service is alive and processing
    queries, not just that port 53 is open.

    Args:
        host: Hostname or IP address to check
        timeout: Query timeout in seconds

    Returns:
        Status string:
        - 'OK' if DNS responded with valid data
        - 'NO_RESPONSE' if connection failed or timed out
        - 'INVALID' if port is open but response is not valid DNS
        - 'ERROR: <details>' for unexpected errors
    """
    logger = logging.getLogger(__name__)

    try:
        # Build a minimal DNS query for "." (root) type NS
        # This is the simplest query any DNS server should answer
        import random
        txn_id = random.randint(0, 65535)

        # DNS header: ID, flags (standard query, recursion desired), 1 question, 0 answers
        header = struct.pack('!HHHHHH',
                             txn_id,    # Transaction ID
                             0x0100,    # Flags: standard query, recursion desired
                             1,         # Questions: 1
                             0,         # Answer RRs: 0
                             0,         # Authority RRs: 0
                             0)         # Additional RRs: 0

        # Question: "." (root) type NS (2) class IN (1)
        # Root domain is encoded as a single zero-length label
        question = b'\x00' + struct.pack('!HH', 2, 1)  # type=NS, class=IN

        dns_query = header + question

        # Send via UDP on port 53
        logger.debug(f"Checking DNS service on {host}:53")
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        sock.settimeout(timeout)

        try:
            sock.sendto(dns_query, (host, 53))
            resp, addr = sock.recvfrom(4096)

            if len(resp) < 12:
                logger.debug(f"DNS {host}: response too short ({len(resp)} bytes)")
                return 'INVALID'

            # Parse response header
            resp_id = struct.unpack('!H', resp[:2])[0]
            resp_flags = struct.unpack('!H', resp[2:4])[0]

            # Verify transaction ID matches
            if resp_id != txn_id:
                logger.debug(f"DNS {host}: transaction ID mismatch (sent {txn_id}, got {resp_id})")
                return 'INVALID'

            # Check QR bit (bit 15) - should be 1 for response
            if not (resp_flags & 0x8000):
                logger.debug(f"DNS {host}: QR bit not set (not a response)")
                return 'INVALID'

            logger.debug(f"DNS {host}: valid response received ({len(resp)} bytes, flags=0x{resp_flags:04x})")
            return 'OK'

        finally:
            sock.close()

    except socket.timeout:
        logger.debug(f"DNS {host}: query timed out")
        return 'NO_RESPONSE'
    except OSError as e:
        logger.debug(f"DNS {host}: connection error: {e}")
        return 'NO_RESPONSE'
    except Exception as e:
        logger.debug(f"DNS {host}: unexpected error: {e}")
        return f'ERROR: {e}'


def check_ldap_service(host: str, timeout: int = 5) -> str:
    """
    Check if LDAP service is responding on port 389 by sending an anonymous bind
    request and verifying a valid LDAP BindResponse is returned.

    Args:
        host: Hostname or IP address to check
        timeout: Connection timeout in seconds

    Returns:
        'OK' if LDAP responded with valid data, 'NO_RESPONSE' if unreachable, 'ERROR: ...' otherwise
    """
    logger = logging.getLogger(__name__)

    try:
        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        sock.connect((host, 389))

        # Anonymous LDAP Simple Bind Request (RFC 4511)
        bind_request = bytes([
            0x30, 0x0c,              # SEQUENCE (12 bytes)
            0x02, 0x01, 0x01,        # INTEGER MessageID = 1
            0x60, 0x07,              # APPLICATION 0 (BindRequest) (7 bytes)
            0x02, 0x01, 0x03,        # INTEGER version = 3
            0x04, 0x00,              # OCTET STRING name = ""
            0x80, 0x00               # CONTEXT 0 (simple auth) = ""
        ])

        sock.sendall(bind_request)
        resp = sock.recv(4096)
        sock.close()

        if resp and len(resp) > 5:
            logger.debug(f"LDAP {host}: responded with {len(resp)} bytes")
            return 'OK'

        logger.debug(f"LDAP {host}: empty or short response")
        return 'NO_RESPONSE'

    except socket.timeout:
        logger.debug(f"LDAP {host}: connection timed out")
        return 'NO_RESPONSE'
    except ConnectionRefusedError:
        logger.debug(f"LDAP {host}: connection refused (port 389 closed)")
        return 'NO_RESPONSE'
    except OSError as e:
        logger.debug(f"LDAP {host}: connection error: {e}")
        return 'NO_RESPONSE'
    except Exception as e:
        logger.debug(f"LDAP {host}: unexpected error: {e}")
        return f'ERROR: {e}'


def check_ldaps_service(host: str, timeout: int = 5) -> tuple[str, Optional[str]]:
    """
    Check if LDAPS service is responding on port 636 by performing a TLS handshake
    and extracting the certificate expiration date.

    Args:
        host: Hostname or IP address to check
        timeout: Connection timeout in seconds

    Returns:
        Tuple of (status, cert_expiry_str):
        - status: 'OK', 'NO_RESPONSE', or 'ERROR: ...'
        - cert_expiry_str: Certificate expiry date as 'YYYY-MM-DD' string, or None
    """
    logger = logging.getLogger(__name__)

    try:
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        sock.settimeout(timeout)
        ssl_sock = ctx.wrap_socket(sock, server_hostname=host)
        ssl_sock.connect((host, 636))

        # Extract certificate expiry from DER
        cert_expiry = None
        try:
            cert_der = ssl_sock.getpeercert(binary_form=True)
            if cert_der:
                cert_expiry = _parse_cert_expiry_from_der(cert_der)
        except Exception as e:
            logger.debug(f"LDAPS {host}: cert parsing error: {e}")

        ssl_sock.close()
        logger.debug(f"LDAPS {host}: TLS handshake OK, cert expiry: {cert_expiry}")
        return 'OK', cert_expiry

    except socket.timeout:
        logger.debug(f"LDAPS {host}: connection timed out")
        return 'NO_RESPONSE', None
    except ConnectionRefusedError:
        logger.debug(f"LDAPS {host}: connection refused (port 636 closed)")
        return 'NO_RESPONSE', None
    except ssl.SSLError as e:
        logger.debug(f"LDAPS {host}: SSL error: {e}")
        return 'NO_RESPONSE', None
    except OSError as e:
        logger.debug(f"LDAPS {host}: connection error: {e}")
        return 'NO_RESPONSE', None
    except Exception as e:
        logger.debug(f"LDAPS {host}: unexpected error: {e}")
        return f'ERROR: {e}', None


def _parse_cert_expiry_from_der(der_bytes: bytes) -> Optional[str]:
    """
    Parse certificate expiry date from DER-encoded certificate bytes.
    Walks the ASN.1 structure to find the Validity sequence and extract notAfter.

    Returns expiry as 'YYYY-MM-DD' string, or None if parsing fails.
    """
    try:
        def _read_tag_len(data, offset):
            """Read ASN.1 tag and length, return (tag, length, header_size)."""
            if offset >= len(data):
                return None, 0, 0
            tag = data[offset]
            if offset + 1 >= len(data):
                return tag, 0, 1
            length_byte = data[offset + 1]
            if length_byte < 0x80:
                return tag, length_byte, 2
            num_bytes = length_byte & 0x7f
            if offset + 2 + num_bytes > len(data):
                return tag, 0, 2
            length = int.from_bytes(data[offset + 2:offset + 2 + num_bytes], 'big')
            return tag, length, 2 + num_bytes

        def _parse_time(data, offset):
            """Parse a UTCTime or GeneralizedTime at offset, return 'YYYY-MM-DD' or None."""
            tag, length, hdr = _read_tag_len(data, offset)
            if tag not in (0x17, 0x18) or length == 0:
                return None
            time_str = data[offset + hdr:offset + hdr + length].decode('ascii')
            if tag == 0x17:  # UTCTime: YYMMDDHHMMSSZ
                year = int(time_str[:2])
                year += 2000 if year < 50 else 1900
                return f"{year:04d}-{int(time_str[2:4]):02d}-{int(time_str[4:6]):02d}"
            else:  # GeneralizedTime: YYYYMMDDHHMMSSZ
                return f"{int(time_str[:4]):04d}-{int(time_str[4:6]):02d}-{int(time_str[6:8]):02d}"

        # Certificate ::= SEQUENCE { tbsCertificate, signatureAlgorithm, signature }
        tag, cert_len, hdr = _read_tag_len(der_bytes, 0)
        if tag != 0x30:
            return None

        # TBSCertificate ::= SEQUENCE { ... }
        tbs_offset = hdr
        tag, tbs_len, tbs_hdr = _read_tag_len(der_bytes, tbs_offset)
        if tag != 0x30:
            return None

        # Walk through TBSCertificate fields to find Validity (5th field typically)
        pos = tbs_offset + tbs_hdr
        tbs_end = tbs_offset + tbs_hdr + tbs_len

        # Skip: version [0] EXPLICIT (optional), serialNumber, signature, issuer
        fields_skipped = 0
        while pos < tbs_end and fields_skipped < 4:
            tag, length, field_hdr = _read_tag_len(der_bytes, pos)
            if tag is None:
                break
            pos += field_hdr + length
            fields_skipped += 1

        # Now pos should be at Validity ::= SEQUENCE { notBefore, notAfter }
        if pos >= tbs_end:
            return None
        tag, val_len, val_hdr = _read_tag_len(der_bytes, pos)
        if tag != 0x30:
            return None

        # Skip notBefore, read notAfter
        nb_offset = pos + val_hdr
        nb_tag, nb_len, nb_hdr = _read_tag_len(der_bytes, nb_offset)
        na_offset = nb_offset + nb_hdr + nb_len
        return _parse_time(der_bytes, na_offset)

    except Exception:
        return None


def parse_server_file(file_path: Path) -> List[str]:
    """
    Auto-detect format and parse server list from TXT or CSV files.

    Implements server list parsing as required by:
    - Requirements 7.1: Parse TXT files with each line as NTP server hostname or IP address
    - Requirements 7.2: Parse CSV files expecting a "server" column header
    - Requirements 7.3: Ignore additional columns and process only server column values
    - Requirements 7.4: Terminate with error if file cannot be read
    - Requirements 7.5: Skip empty lines and whitespace-only entries

    Args:
        file_path: Path to server list file (.txt or .csv)

    Returns:
        List of NTP server hostnames/IPs from the file

    Raises:
        SystemExit: If file cannot be read (terminates with error message)
    """
    logger = logging.getLogger(__name__)

    try:
        if not file_path.exists():
            logger.error(f"Server list file not found: {file_path}")
            sys.exit(1)

        if not file_path.is_file():
            logger.error(f"Server list path is not a file: {file_path}")
            sys.exit(1)

        # Auto-detect format based on file extension
        if file_path.suffix.lower() == '.csv':
            logger.info(f"Parsing CSV server list: {file_path}")
            return parse_csv_file(file_path)
        elif file_path.suffix.lower() == '.txt':
            logger.info(f"Parsing TXT server list: {file_path}")
            return parse_txt_file(file_path)
        else:
            # Default to TXT format for unknown extensions
            logger.info(f"Unknown file extension, treating as TXT format: {file_path}")
            return parse_txt_file(file_path)

    except Exception as e:
        logger.error(f"Cannot read server list file {file_path}: {e}")
        sys.exit(1)


def parse_txt_file(file_path: Path) -> List[str]:
    """
    Parse line-delimited server list from TXT file.

    Args:
        file_path: Path to TXT file

    Returns:
        List of NTP server hostnames/IPs, with empty lines and whitespace skipped
    """
    logger = logging.getLogger(__name__)
    servers = []

    try:
        logger.debug(f"Opening TXT file for parsing: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                # Skip empty lines and whitespace-only entries
                server = line.strip()
                if server:
                    servers.append(server)
                    logger.debug(f"Line {line_num}: Added server '{server}'")
                else:
                    logger.debug(f"Line {line_num}: Skipped empty line")

        logger.info(f"Parsed {len(servers)} servers from TXT file")
        logger.debug(f"Server list: {', '.join(servers)}")
        return servers

    except Exception as e:
        logger.error(f"Error reading TXT file {file_path}: {e}")
        raise


def parse_csv_file(file_path: Path) -> List[str]:
    """
    Parse CSV file expecting a "server" column header.

    Args:
        file_path: Path to CSV file

    Returns:
        List of NTP server hostnames/IPs from server column, with empty entries skipped
    """
    logger = logging.getLogger(__name__)
    servers = []

    try:
        logger.debug(f"Opening CSV file for parsing: {file_path}")
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)

            # Log available columns for debugging
            logger.debug(f"CSV columns found: {reader.fieldnames}")

            # Check if server column exists
            if 'server' not in reader.fieldnames:
                logger.error(f"CSV file missing required 'server' column. Found columns: {reader.fieldnames}")
                raise Exception("Missing 'server' column in CSV file")

            # Log additional columns that will be ignored
            additional_columns = [col for col in reader.fieldnames if col != 'server']
            if additional_columns:
                logger.debug(f"Ignoring additional CSV columns: {', '.join(additional_columns)}")

            for row_num, row in enumerate(reader, 2):  # Start at 2 since row 1 is header
                # Process only the server column values, ignore additional columns
                server = row.get('server', '').strip()
                if server:
                    servers.append(server)
                    logger.debug(f"Row {row_num}: Added server '{server}'")
                else:
                    logger.debug(f"Row {row_num}: Skipped empty server entry")

        logger.info(f"Parsed {len(servers)} servers from CSV file")
        logger.debug(f"Server list: {', '.join(servers)}")
        return servers

    except Exception as e:
        logger.error(f"Error reading CSV file {file_path}: {e}")
        raise


def parse_additional_servers_csv(file_path: Path) -> List[tuple[str, str, int, int, str, str, int]]:
    """
    Parse additional servers CSV file with extended tracking columns.
    Supports backward compatibility with 2-column, 3-column, 6-column, and 7-column formats.

    Args:
        file_path: Path to CSV file

    Returns:
        List of tuples (server, short_name, failure_count, missed, failed, dlist, time)
        - server: NTP server hostname or IP
        - short_name: Display name for the server
        - failure_count: Consecutive failure count (existing)
        - missed: Total connection error count (never resets)
        - failed: Status string ('failed' if missed >= 10, 'ok' otherwise)
        - dlist: Distribution list email address
        - time: Consecutive variance threshold violations (resets when within threshold)
    """
    logger = logging.getLogger(__name__)
    servers = []

    try:
        logger.debug(f"Opening additional servers CSV file for parsing: {file_path}")

        # Read the CSV file and detect columns
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            columns = reader.fieldnames

            if not columns or 'server' not in columns:
                logger.error(f"CSV file missing required 'server' column. Found columns: {columns}")
                raise Exception("Missing 'server' column in CSV file")

            logger.debug(f"Detected CSV columns: {columns}")

            # Determine which columns are present
            has_short_name = 'short_name' in columns or 'location' in columns
            has_failure_count = 'failure_count' in columns
            has_missed = 'missed' in columns
            has_failed = 'failed' in columns
            has_dlist = 'dlist' in columns
            has_time = 'time' in columns

            # Log missing columns and defaults
            missing_columns = []
            if not has_missed:
                missing_columns.append('missed')
            if not has_failed:
                missing_columns.append('failed')
            if not has_dlist:
                missing_columns.append('dlist')
            if not has_time:
                missing_columns.append('time')

            if missing_columns:
                logger.info(f"Columns not found: {', '.join(missing_columns)}")
                logger.info(f"Will apply defaults: missed=0, failed='ok', dlist='none', time=0")

            # Determine name column
            name_column = 'short_name' if 'short_name' in columns else ('location' if 'location' in columns else None)

            # Parse each row
            for row_num, row in enumerate(reader, 2):  # Start at 2 since row 1 is header
                # Process server column (required)
                server = row.get('server', '').strip()
                if not server:
                    logger.warning(f"Row {row_num}: Skipped empty server entry")
                    continue

                # Process short_name (optional, defaults to server name)
                short_name = ''
                if name_column:
                    short_name = row.get(name_column, '').strip()
                if not short_name:
                    short_name = server

                # Process failure_count (optional, defaults to 0)
                failure_count = 0
                if has_failure_count:
                    failure_count_str = row.get('failure_count', '').strip()
                    if failure_count_str:
                        try:
                            failure_count = int(failure_count_str)
                            if failure_count < 0:
                                logger.warning(f"Row {row_num}: Negative failure_count {failure_count} for server {server}, replacing with 0")
                                failure_count = 0
                        except ValueError:
                            logger.warning(f"Row {row_num}: Invalid failure_count '{failure_count_str}' for server {server}, replacing with 0")
                            failure_count = 0

                # Process missed (optional, defaults to 0)
                missed = 0
                if has_missed:
                    missed_str = row.get('missed', '').strip()
                    if missed_str:
                        try:
                            missed = int(missed_str)
                            if missed < 0:
                                logger.warning(f"Row {row_num}: Negative missed value {missed} for server {server}, replacing with 0")
                                missed = 0
                        except ValueError:
                            logger.warning(f"Row {row_num}: Invalid missed value '{missed_str}' for server {server}, replacing with 0")
                            missed = 0

                # Process failed (optional, defaults to 'ok')
                failed = 'ok'
                if has_failed:
                    failed_str = row.get('failed', '').strip().lower()
                    if failed_str in ['ok', 'failed']:
                        failed = failed_str
                    elif failed_str:
                        logger.warning(f"Row {row_num}: Invalid failed status '{failed_str}' for server {server}, replacing with 'ok'")

                # Process dlist (optional, defaults to 'none')
                dlist = 'none'
                if has_dlist:
                    dlist_str = row.get('dlist', '').strip()
                    if dlist_str:
                        dlist = dlist_str

                # Process time (optional, defaults to 0)
                time = 0
                if has_time:
                    time_str = row.get('time', '').strip()
                    if time_str:
                        try:
                            time = int(time_str)
                            if time < 0:
                                logger.warning(f"Row {row_num}: Negative time value {time} for server {server}, replacing with 0")
                                time = 0
                        except ValueError:
                            logger.warning(f"Row {row_num}: Invalid time value '{time_str}' for server {server}, replacing with 0")
                            time = 0

                servers.append((server, short_name, failure_count, missed, failed, dlist, time))
                logger.debug(f"Row {row_num}: Parsed server '{server}' (short_name='{short_name}', failure_count={failure_count}, missed={missed}, failed='{failed}', dlist='{dlist}', time={time})")

        logger.info(f"Parsed {len(servers)} servers from CSV file")
        return servers

    except Exception as e:
        logger.error(f"Error reading additional servers CSV file {file_path}: {e}")
        raise


def write_additional_servers_csv(file_path: Path, servers: List[tuple[str, str, int, int, str, str, int]]) -> None:
    """
    Write updated server data with all tracking columns to CSV file using atomic write operation.

    Implements CSV writing as required by:
    - Requirements 5.1: Write header row with all 7 columns in correct order
    - Requirements 5.2: Write one data row per server with all 7 values
    - Requirements 5.3: Use UTF-8 encoding
    - Requirements 5.4: Use proper CSV formatting with quoted fields
    - Requirements 5.5: Log each row being written
    - Requirements 6.1: Create temporary file with .tmp extension
    - Requirements 6.2: Rename temporary file to target filename
    - Requirements 6.3: Log success or failure with specific error details
    - Requirements 6.4: Delete temporary file on failure
    - Requirements 6.5: Ensure original file unchanged on failure
    - Requirements 6.6: Make write failures non-fatal (log and continue)

    Args:
        file_path: Path to CSV file
        servers: List of tuples (server, short_name, failure_count, missed, failed, dlist, time)

    Raises:
        Exception: If file write operation fails (logged but non-fatal)
    """
    logger = logging.getLogger(__name__)
    temp_path = None

    try:
        # Create temporary file with .tmp extension (Requirement 6.1)
        temp_path = file_path.with_suffix(file_path.suffix + '.tmp')

        logger.debug(f"Writing to temporary file: {temp_path}")

        with open(temp_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)

            # Write header row with all 7 columns in correct order (Requirement 5.1)
            writer.writerow(['server', 'short_name', 'failure_count', 'missed', 'failed', 'dlist', 'time'])
            logger.debug("Wrote header row: ['server', 'short_name', 'failure_count', 'missed', 'failed', 'dlist', 'time']")

            # Write data rows (Requirements 5.2, 5.5)
            for server, short_name, failure_count, missed, failed, dlist, time in servers:
                writer.writerow([server, short_name, failure_count, missed, failed, dlist, time])
                logger.debug(f"Wrote row: server={server}, short_name={short_name}, failure_count={failure_count}, "
                           f"missed={missed}, failed={failed}, dlist={dlist}, time={time}")

        # Atomic rename operation (Requirement 6.2)
        temp_path.replace(file_path)
        logger.info(f"Successfully wrote CSV file with all columns to {file_path}")

    except Exception as e:
        # Non-fatal error: log and continue (Requirement 6.6)
        logger.error(f"Failed to write CSV file to {file_path}: {e}")
        logger.debug(f"Error details: {type(e).__name__}: {str(e)}")

        # Delete temporary file on failure (Requirement 6.4)
        if temp_path and temp_path.exists():
            try:
                temp_path.unlink()
                logger.debug(f"Cleaned up temporary file: {temp_path}")
            except Exception as cleanup_error:
                logger.warning(f"Failed to clean up temporary file {temp_path}: {cleanup_error}")

        logger.debug("Continuing program execution despite CSV write failure")
        # Don't raise - failure tracking is supplementary functionality


def get_all_domain_ips(domain: str, timeout: int = 10) -> List[str]:
    """
    Get all A records (IP addresses) for the specified domain.

    Args:
        domain: Domain to query for A records (e.g., 'tgna.tegna.com')
        timeout: DNS query timeout in seconds

    Returns:
        List of IP addresses found for the domain
    """
    logger = logging.getLogger(__name__)
    ip_addresses = []

    try:
        # Configure DNS resolver with timeout
        resolver = dns.resolver.Resolver()
        resolver.timeout = timeout
        resolver.lifetime = timeout

        logger.info(f"Querying all A records for domain: {domain}")

        # Get all A records for the domain (try UDP first, fall back to TCP for large record sets)
        try:
            answers = resolver.resolve(domain, 'A')

            for rdata in answers:
                ip = str(rdata)
                ip_addresses.append(ip)
                logger.debug(f"Found A record: {domain} -> {ip}")

        except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer) as e:
            logger.warning(f"No A records found for domain {domain}: {e}")
            return []
        except (dns.resolver.Timeout, dns.exception.Timeout) as e:
            logger.warning(f"DNS UDP query timeout for {domain}, trying TCP")
            try:
                answers = resolver.resolve(domain, 'A', tcp=True)
                for rdata in answers:
                    ip = str(rdata)
                    ip_addresses.append(ip)
                    logger.debug(f"Found A record (TCP): {domain} -> {ip}")
            except Exception as e2:
                logger.warning(f"DNS TCP query also failed for {domain}: {e2}")
                return []
        except Exception as e:
            logger.error(f"DNS query error for domain {domain}: {e}")
            return []

        if ip_addresses:
            logger.info(f"Found {len(ip_addresses)} A records for {domain}: {', '.join(ip_addresses)}")
        else:
            logger.warning(f"No A records found for domain: {domain}")

        return ip_addresses

    except Exception as e:
        logger.error(f"Failed to query A records for domain {domain}: {e}")
        return []


def discover_ntp_servers_in_domain(domain: str, timeout: int = 10) -> List[str]:
    """
    Discover NTP servers in the specified domain by getting all A records and using them as potential NTP servers.

    Args:
        domain: Domain to search for NTP servers (e.g., 'tgna.tegna.com')
        timeout: DNS query timeout in seconds

    Returns:
        List of IP addresses from A records that can be used as NTP servers
    """
    logger = logging.getLogger(__name__)

    # Get all A records for the domain
    ip_addresses = get_all_domain_ips(domain, timeout)

    if not ip_addresses:
        logger.warning(f"No A records found for domain {domain}")
        return []

    logger.info(f"Using all {len(ip_addresses)} A records from {domain} as potential NTP servers")

    # Return all IP addresses as potential NTP servers
    # The NTP query process will determine which ones actually provide NTP service
    return ip_addresses


def perform_reverse_dns_lookup(ip_address: str, domain_suffix: str = "tgna.tegna.com", timeout: int = 10) -> tuple[Optional[str], Optional[str]]:
    """
    Perform reverse DNS lookup on an IP address and extract hostname information.

    Args:
        ip_address: IP address to perform reverse lookup on
        domain_suffix: Domain suffix to remove for short names (e.g., "tgna.tegna.com")
        timeout: DNS query timeout in seconds

    Returns:
        Tuple of (full_hostname, short_name)
        - full_hostname: Complete hostname from reverse DNS
        - short_name: Hostname with domain suffix removed (if applicable)
    """
    logger = logging.getLogger(__name__)

    try:
        # Configure DNS resolver with timeout
        resolver = dns.resolver.Resolver()
        resolver.timeout = timeout
        resolver.lifetime = timeout

        # Create reverse DNS query
        reverse_name = dns.reversename.from_address(ip_address)

        logger.debug(f"Performing reverse DNS lookup for: {ip_address}")

        # Perform reverse DNS lookup
        answers = resolver.resolve(reverse_name, 'PTR')

        if answers:
            full_hostname = str(answers[0]).rstrip('.')  # Remove trailing dot
            logger.debug(f"Reverse DNS successful: {ip_address} -> {full_hostname}")

            # Extract short name by removing configured domain suffix
            short_name = full_hostname
            domain_suffix_with_dot = f'.{domain_suffix}'
            if full_hostname.endswith(domain_suffix_with_dot):
                short_name = full_hostname[:-len(domain_suffix_with_dot)]
                logger.debug(f"Extracted short name: {short_name}")

            return full_hostname, short_name
        else:
            logger.debug(f"No PTR record found for: {ip_address}")
            return None, None

    except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer):
        logger.debug(f"No reverse DNS record found for: {ip_address}")
        return None, None
    except dns.resolver.Timeout:
        logger.debug(f"Reverse DNS timeout for: {ip_address}")
        return None, None
    except Exception as e:
        logger.debug(f"Reverse DNS error for {ip_address}: {e}")
        return None, None


def resolve_hostname_with_fallback(hostname: str, timeout: int = 10) -> tuple[str, Optional[str]]:
    """
    Attempt DNS resolution for hostname with fallback to original hostname.

    Args:
        hostname: Hostname or IP address to resolve
        timeout: DNS resolution timeout in seconds

    Returns:
        Tuple of (hostname_to_use, resolved_ip_or_none)
    """
    logger = logging.getLogger(__name__)

    # If it's already an IP address, return as-is
    try:
        socket.inet_aton(hostname)
        logger.debug(f"Hostname {hostname} is already an IP address")
        return hostname, hostname
    except socket.error:
        pass  # Not an IP address, continue with DNS resolution

    try:
        # Configure DNS resolver with timeout
        resolver = dns.resolver.Resolver()
        resolver.timeout = timeout
        resolver.lifetime = timeout

        logger.debug(f"Attempting DNS resolution for: {hostname}")
        answers = resolver.resolve(hostname, 'A')

        if answers:
            resolved_ip = str(answers[0])
            logger.debug(f"DNS resolution successful: {hostname} -> {resolved_ip}")
            return hostname, resolved_ip
        else:
            logger.debug(f"DNS resolution returned no answers for: {hostname}")
            return hostname, None

    except (dns.resolver.NXDOMAIN, dns.resolver.NoAnswer, dns.resolver.Timeout) as e:
        logger.debug(f"DNS resolution failed for {hostname}: {e}")
        return hostname, None
    except Exception as e:
        logger.debug(f"DNS resolution error for {hostname}: {e}")
        return hostname, None


def calculate_time_delta(target_time: Optional[datetime], reference_time: datetime) -> Optional[float]:
    """
    Calculate signed time delta between target NTP time and reference NTP time.

    Args:
        target_time: Target NTP server timestamp (None if query failed)
        reference_time: Reference NTP server timestamp

    Returns:
        Delta in seconds (target_time - reference_time), None if target_time is None
        Positive values indicate target is ahead of reference
        Negative values indicate target is behind reference
    """
    if target_time is None:
        return None

    # Calculate delta as target_time minus reference_time
    delta_seconds = (target_time - reference_time).total_seconds()

    return delta_seconds


def process_single_server(server: str, reference_offset: float, reference_query_time: datetime, config: Config, ini_config: dict, is_additional_server: bool = False, custom_short_name: str = None, current_failure_count: int = 0, current_missed: int = 0, current_failed: str = 'ok', current_time: int = 0) -> NTPResult:
    """
    Process a single NTP server with DNS resolution, query, and delta calculation.

    Implements graceful failure handling as required by:
    - Requirements 3.5: Continue processing remaining servers on individual failures
    - Requirements 3.6: Track failure counts for each server
    - Requirements 4.3: Update failure count based on query result
    - Requirements 3.1, 3.2, 9.1, 9.2, 9.3, 9.4, 9.5: Update tracking counters independently

    Args:
        server: NTP server hostname or IP address
        reference_offset: Reference NTP server offset (from batch query)
        reference_query_time: Wall clock time when reference was queried (unused, kept for compatibility)
        config: Configuration object with timeout and format settings
        is_additional_server: True if this is an additional server (excluded from error counts)
        custom_short_name: Custom short name for the server (optional)
        current_failure_count: Current consecutive failure count for this server
        current_missed: Current total connection error count (never resets)
        current_failed: Current failed status ('ok' or 'failed')

    Returns:
        NTPResult with complete query results, status, and updated counters (failure_count, missed, failed).
        Always returns a result object, even for failed queries, to maintain processing continuity.
    """
    logger = logging.getLogger(__name__)
    query_timestamp = datetime.now(timezone.utc)

    # Attempt DNS resolution with fallback - graceful handling of DNS failures
    hostname_to_use, resolved_ip = resolve_hostname_with_fallback(server, config.ntp_timeout)

    # Perform reverse DNS lookup to get hostname information
    full_hostname = None
    short_name = None

    # Use custom short name if provided (for additional servers)
    if custom_short_name:
        short_name = custom_short_name
        logger.debug(f"Using custom short name: {custom_short_name}")

    if resolved_ip:
        full_hostname, dns_short_name = perform_reverse_dns_lookup(resolved_ip, ini_config['default_discovery_domain'], config.ntp_timeout)
        # Only use DNS-derived short name if no custom short name was provided
        if not custom_short_name:
            short_name = dns_short_name

    # Check Kerberos KDC and DNS service status (primary servers only, not additional)
    kerberos_status = None
    dns_status = None
    ldap_status = None
    ldaps_status = None
    ldaps_cert_expiry = None
    if not is_additional_server:
        check_target = resolved_ip if resolved_ip else server
        kerberos_status = check_kerberos_kdc(check_target, timeout=min(config.ntp_timeout, 5))
        dns_status = check_dns_service(check_target, timeout=min(config.ntp_timeout, 5))
        ldap_status = check_ldap_service(check_target, timeout=min(config.ntp_timeout, 5))
        ldaps_status, ldaps_cert_expiry = check_ldaps_service(check_target, timeout=min(config.ntp_timeout, 5))

    try:
        # Query the NTP server
        logger.debug(f"Querying NTP server: {server} (using {hostname_to_use})")
        ntp_response = query_ntp_server(hostname_to_use, config.ntp_timeout)

        # Calculate delta using NTP offsets
        # The offset from each server tells us how much that server differs from our local clock
        # To find the difference between two servers, we subtract their offsets
        # Delta = target_offset - reference_offset
        # This gives us the time difference between the two NTP servers
        delta_seconds = ntp_response.offset_seconds - reference_offset
        delta_formatted = format_delta_value(delta_seconds, config.format_type)
        logger.debug(f"Delta calculated: {delta_seconds:.6f}s (target offset: {ntp_response.offset_seconds:.6f}s, reference offset: {reference_offset:.6f}s)")

        # Determine final status based on validation
        status, error_message = validate_ntp_response(ntp_response)

        # Update failure count based on status
        updated_failure_count = update_failure_count(server, current_failure_count, status)

        # Update tracking counters (missed, failed status, and time)
        # Get variance threshold from config
        variance_threshold_ms = ini_config.get('variance_threshold_ms', 33.0)
        updated_missed, updated_failed, updated_time = update_tracking_counters(
            server, current_missed, current_failed, status, delta_seconds, variance_threshold_ms, current_time
        )

        return NTPResult(
            timestamp_utc=query_timestamp,
            ntp_server=server,
            ntp_server_ip=resolved_ip,
            hostname=full_hostname,
            short_name=short_name,
            ntp_time_utc=ntp_response.timestamp_utc,
            query_rtt_ms=ntp_response.query_rtt_ms,
            stratum=ntp_response.stratum,
            root_delay_ms=ntp_response.root_delay_ms,
            root_dispersion_ms=ntp_response.root_dispersion_ms,
            delta_seconds=delta_seconds,
            delta_formatted=delta_formatted,
            status=status,
            error_message=error_message,
            is_additional_server=is_additional_server,
            failure_count=updated_failure_count,
            missed=updated_missed,
            failed=updated_failed,
            time=updated_time,
            leap_indicator=ntp_response.leap_indicator,
            precision=ntp_response.precision,
            ref_id=ntp_response.ref_id,
            ref_time=ntp_response.ref_time,
            poll_interval=ntp_response.poll_interval,
            kerberos_status=kerberos_status,
            dns_status=dns_status,
            ldap_status=ldap_status,
            ldaps_status=ldaps_status,
            ldaps_cert_expiry=ldaps_cert_expiry
        )

    except Exception as e:
        # Graceful failure handling - classify error and return result object
        # This ensures processing continues for remaining servers
        status, error_message = handle_ntp_query_error(server, e, config.ntp_timeout)

        logger.debug(f"Server {server} failed with status {status.value}: {error_message}")

        # Update failure count based on error status
        updated_failure_count = update_failure_count(server, current_failure_count, status)

        # Update tracking counters (missed, failed status, and time)
        # For errors, delta_seconds is None, so time counter won't be updated
        variance_threshold_ms = ini_config.get('variance_threshold_ms', 33.0)
        updated_missed, updated_failed, updated_time = update_tracking_counters(
            server, current_missed, current_failed, status, None, variance_threshold_ms, current_time
        )

        return NTPResult(
            timestamp_utc=query_timestamp,
            ntp_server=server,
            ntp_server_ip=resolved_ip,
            hostname=full_hostname,
            short_name=short_name,
            ntp_time_utc=None,
            query_rtt_ms=None,
            stratum=None,
            root_delay_ms=None,
            root_dispersion_ms=None,
            delta_seconds=None,  # No delta calculation for failed queries
            delta_formatted=None,  # No formatted delta for failed queries
            status=status,
            error_message=error_message,
            is_additional_server=is_additional_server,
            failure_count=updated_failure_count,
            missed=updated_missed,
            failed=updated_failed,
            time=updated_time,
            kerberos_status=kerberos_status,
            dns_status=dns_status,
            ldap_status=ldap_status,
            ldaps_status=ldaps_status,
            ldaps_cert_expiry=ldaps_cert_expiry
        )


def process_servers_parallel(server_list, reference_server: str, config: Config, ini_config: dict, is_additional_servers: bool = False, failure_counts: dict = None, missed_counts: dict = None, failed_statuses: dict = None, time_counts: dict = None) -> List[NTPResult]:
    """
    Process multiple NTP servers concurrently using thread pool.

    Implements graceful failure handling as required by:
    - Requirements 3.5: Continue processing remaining servers on individual failures
    - Requirements 3.6: Pass failure counts to individual server processing
    - Requirements 3.1, 3.2, 9.1, 9.2, 9.3, 9.4, 9.5: Pass tracking counters to individual server processing

    Args:
        server_list: List of NTP server hostnames/IPs (strings) or tuples (server, short_name) for additional servers
        reference_server: Reference NTP server hostname for batch reference query
        config: Configuration object with parallel limits and timeout settings
        is_additional_servers: True if processing additional servers (excludes from error counts)
        failure_counts: Dictionary mapping server hostname to current failure count (optional)
        missed_counts: Dictionary mapping server hostname to current missed count (optional)
        failed_statuses: Dictionary mapping server hostname to current failed status (optional)

    Returns:
        List of NTPResult objects with query results for all servers.
        Results are collected regardless of individual server errors.
        Processing state is maintained across failures.
    """
    logger = logging.getLogger(__name__)
    results = []

    if not server_list:
        logger.warning("No servers to process")
        return results

    # Query reference server once at the start of batch processing
    logger.info(f"Querying reference server once for batch: {reference_server}")
    reference_query_time = datetime.now(timezone.utc)
    try:
        reference_response = query_ntp_server(reference_server, config.ntp_timeout)
        reference_offset = reference_response.offset_seconds
        logger.info(f"Reference offset captured: {reference_offset:.6f}s (stratum {reference_response.stratum})")
    except Exception as e:
        logger.error(f"Failed to query reference server {reference_server}: {e}")
        logger.error("Cannot proceed without reference time")
        return results

    # Use configured parallel limit, default to 10 if not specified
    max_workers = config.parallel_limit if config.parallel_limit > 0 else 10

    logger.info(f"Processing {len(server_list)} servers with {max_workers} concurrent workers")

    # Track processing statistics for graceful failure handling
    total_servers = len(server_list)
    completed_servers = 0
    successful_servers = 0
    failed_servers = 0

    # Create thread pool and submit all server processing tasks
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # Submit all tasks - continue processing even if some fail
        future_to_server = {}

        # Initialize failure_counts dict if not provided
        if failure_counts is None:
            failure_counts = {}

        # Initialize missed_counts dict if not provided
        if missed_counts is None:
            missed_counts = {}

        # Initialize failed_statuses dict if not provided
        if failed_statuses is None:
            failed_statuses = {}

        # Initialize time_counts dict if not provided
        if time_counts is None:
            time_counts = {}

        for item in server_list:
            if is_additional_servers and isinstance(item, tuple):
                # Additional servers: (server, short_name)
                server, custom_short_name = item
                current_failure_count = failure_counts.get(server, 0)
                current_missed = missed_counts.get(server, 0)
                current_failed = failed_statuses.get(server, 'ok')
                current_time = time_counts.get(server, 0)
                future = executor.submit(process_single_server, server, reference_offset, reference_query_time, config, ini_config, is_additional_servers, custom_short_name, current_failure_count, current_missed, current_failed, current_time)
            else:
                # Regular servers: just server string
                server = item
                current_failure_count = failure_counts.get(server, 0)
                current_missed = missed_counts.get(server, 0)
                current_failed = failed_statuses.get(server, 'ok')
                current_time = time_counts.get(server, 0)
                future = executor.submit(process_single_server, server, reference_offset, reference_query_time, config, ini_config, is_additional_servers, None, current_failure_count, current_missed, current_failed, current_time)

            future_to_server[future] = server if isinstance(item, str) else item[0]

        # Collect results as they complete - maintain processing state across failures
        for future in as_completed(future_to_server):
            server = future_to_server[future]
            completed_servers += 1

            try:
                result = future.result()
                results.append(result)

                # Track success/failure statistics
                if result.status == NTPStatus.OK:
                    successful_servers += 1
                    logger.debug(f"Successfully processed server {completed_servers}/{total_servers}: {server}")
                else:
                    failed_servers += 1
                    logger.debug(f"Server {completed_servers}/{total_servers} failed with status {result.status.value}: {server}")
                    if config.verbose and result.error_message:
                        logger.debug(f"  Error details: {result.error_message}")

            except Exception as e:
                # This should not happen since process_single_server handles all exceptions
                # But we implement graceful handling even for unexpected failures
                failed_servers += 1
                logger.error(f"Unexpected error processing server {completed_servers}/{total_servers} {server}: {e}")

                # Create error result for unexpected failures - collect all results regardless of errors
                error_result = NTPResult(
                    timestamp_utc=datetime.now(timezone.utc),
                    ntp_server=server,
                    ntp_server_ip=None,
                    hostname=None,
                    short_name=None,
                    ntp_time_utc=None,
                    query_rtt_ms=None,
                    stratum=None,
                    root_delay_ms=None,
                    root_dispersion_ms=None,
                    delta_seconds=None,
                    delta_formatted=None,
                    status=NTPStatus.ERROR,
                    error_message=f"Unexpected processing error: {e}",
                    is_additional_server=is_additional_servers
                )
                results.append(error_result)

    # Log final processing statistics - maintain processing state across failures
    logger.info(f"Completed processing {completed_servers}/{total_servers} servers")
    logger.info(f"Success: {successful_servers}, Failed: {failed_servers}")

    if failed_servers > 0:
        logger.info(f"Graceful failure handling: Collected results for all {total_servers} servers despite {failed_servers} failures")

    return results


def load_configuration(config_file: str = "ntp_monitor.ini") -> dict:
    """
    Load configuration from INI file with fallback defaults.

    Args:
        config_file: Path to configuration file

    Returns:
        Dictionary containing configuration values
    """
    logger = logging.getLogger(__name__)

    # Default configuration values (used when no INI file exists)
    defaults = {
        'default_reference_server': 'time.cloudflare.com',
        'default_discovery_domain': 'pool.ntp.org',
        'fallback_servers': ['time.google.com', 'time.cloudflare.com', 'time.aws.com', 'time.windows.com'],
        'default_format': 'milliseconds',
        'default_parallel_limit': 10,
        'default_timeout': 30,
        'output_directory': '.\\Reports',
        'sort_by_variance': True,
        'max_variance_display': 100,
        # Email settings
        'send_email': False,
        'smtp_server': '',
        'smtp_port': 25,
        'smtp_use_tls': False,
        'smtp_username': '',
        'smtp_password': '',
        'from_email': '',
        'to_email': '',
        'variance_threshold_ms': 33,
        'skip_threshold': 10
    }

    config = configparser.ConfigParser()

    try:
        if Path(config_file).exists():
            config.read(config_file)
            logger.debug(f"Loaded configuration from {config_file}")

            # Extract values from INI file
            if 'ntp_settings' in config:
                ntp_section = config['ntp_settings']
                defaults['default_reference_server'] = ntp_section.get('default_reference_server', defaults['default_reference_server'])
                defaults['default_discovery_domain'] = ntp_section.get('default_discovery_domain', defaults['default_discovery_domain'])

                # Parse fallback servers list
                fallback_str = ntp_section.get('fallback_servers', '')
                if fallback_str:
                    defaults['fallback_servers'] = [s.strip() for s in fallback_str.split(',')]

            if 'report_settings' in config:
                report_section = config['report_settings']
                defaults['default_format'] = report_section.get('default_format', defaults['default_format'])
                defaults['default_parallel_limit'] = report_section.getint('default_parallel_limit', defaults['default_parallel_limit'])
                defaults['default_timeout'] = report_section.getint('default_timeout', defaults['default_timeout'])
                defaults['output_directory'] = report_section.get('output_directory', defaults['output_directory'])

            if 'advanced_settings' in config:
                advanced_section = config['advanced_settings']
                defaults['sort_by_variance'] = advanced_section.getboolean('sort_by_variance', defaults['sort_by_variance'])
                defaults['max_variance_display'] = advanced_section.getint('max_variance_display', defaults['max_variance_display'])
                defaults['skip_threshold'] = advanced_section.getint('skip_threshold', defaults['skip_threshold'])

            if 'email_settings' in config:
                email_section = config['email_settings']
                defaults['send_email'] = email_section.getboolean('send_email', defaults['send_email'])
                defaults['smtp_server'] = email_section.get('smtp_server', defaults['smtp_server'])
                defaults['smtp_port'] = email_section.getint('smtp_port', defaults['smtp_port'])
                defaults['smtp_use_tls'] = email_section.getboolean('smtp_use_tls', defaults['smtp_use_tls'])
                defaults['smtp_username'] = email_section.get('smtp_username', defaults['smtp_username'])
                defaults['smtp_password'] = email_section.get('smtp_password', defaults['smtp_password'])
                defaults['from_email'] = email_section.get('from_email', defaults['from_email'])
                defaults['to_email'] = email_section.get('to_email', defaults['to_email'])
                defaults['variance_threshold_ms'] = email_section.getfloat('variance_threshold_ms', defaults['variance_threshold_ms'])

        else:
            logger.debug(f"Configuration file {config_file} not found, using defaults")

    except Exception as e:
        logger.warning(f"Error reading configuration file {config_file}: {e}")
        logger.debug("Using default configuration values")

    return defaults


def setup_logging(verbose: bool = False) -> None:
    """
    Configure logging system based on verbosity level.

    Implements detailed operation logging as required by:
    - Requirements 6.5: Output detailed operation information when verbose logging is enabled

    Args:
        verbose: Enable detailed DEBUG level logging if True, INFO level if False
    """
    log_level = logging.DEBUG if verbose else logging.INFO
    log_format = '%(asctime)s - %(levelname)s - %(message)s'

    # Configure root logger
    logging.basicConfig(
        level=log_level,
        format=log_format,
        handlers=[
            logging.StreamHandler(sys.stdout)
        ]
    )

    # Set specific loggers to appropriate levels
    logger = logging.getLogger(__name__)

    if verbose:
        logger.info("Verbose logging enabled - detailed operation information will be displayed")
        logger.debug("Debug logging level activated")

        # Enable detailed logging for DNS operations
        dns_logger = logging.getLogger('dns')
        dns_logger.setLevel(logging.WARNING)  # Reduce DNS library noise

        # Log configuration details when verbose
        logger.debug("Logging configuration completed")
    else:
        # In non-verbose mode, reduce noise from external libraries
        logging.getLogger('dns').setLevel(logging.WARNING)
        logging.getLogger('urllib3').setLevel(logging.WARNING)


def parse_arguments() -> Config:
    """
    Parse command line arguments and return configuration.

    Implements comprehensive CLI argument handling as required by:
    - Requirements 1.1: Support reference NTP parameter
    - Requirements 3.2: Support parallel limit specification (default 10)
    - Requirements 3.4: Support timeout specification (default 30 seconds)
    - Requirements 4.2: Support seconds format for delta values
    - Requirements 4.3: Support milliseconds format for delta values

    Returns:
        Config object with validated command line arguments

    Raises:
        SystemExit: If argument validation fails or help is requested
    """
    # Load configuration from INI file
    ini_config = load_configuration()

    parser = argparse.ArgumentParser(
        description='NTP Delta Monitor - Query multiple NTP servers for time synchronization analysis',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Examples:
  Default usage (auto-discover NTP servers from {ini_config['default_discovery_domain']}):
    %(prog)s

  Basic usage with TXT server list:
    %(prog)s -r ntp1.example.com -s servers.txt

  CSV server list with custom output and milliseconds format:
    %(prog)s -r ntp2.example.com -s servers.csv -o report.xlsx -f milliseconds

  High concurrency with custom timeout and verbose logging:
    %(prog)s -r 10.43.9.64 -s ntp_list.txt -p 20 -t 10 -v

  Milliseconds format with custom output file:
    %(prog)s -r 10.176.127.84 -s servers.csv -f milliseconds -o sync_report.xlsx

  Include additional servers for monitoring (excluded from error counts):
    %(prog)s -r ntp1.example.com -s primary_servers.txt --additional-servers monitoring_servers.csv

  Auto-detect external_ntp_servers.csv in current directory:
    %(prog)s
    (Will automatically use external_ntp_servers.csv if present)

Server List File Formats:
  TXT format: One NTP server hostname or IP per line
    Example:
      ntp1.example.com
      ntp2.example.com
      192.168.1.100

  CSV format: Must have 'server' column header
    Example:
      server,location,notes
      ntp1.example.com,datacenter1,primary
      ntp2.example.com,datacenter2,backup

Additional Servers CSV Format:
  CSV format with 'server' and optional 'short_name' or 'location' columns:
    Example with short_name:
      server,short_name
      10.43.9.64,GPS-KING
      10.176.127.84,GPS-KGW
      pool.ntp.org,
      time.google.com,

    Example with location:
      server,location,notes
      pool.ntp.org,NTP Pool,Public NTP server
      10.43.9.64,GPS-KING,TEGNA Real GPS at KING
      time.google.com,time.google.com,GOOGLE domain time default

  - First column 'server': NTP server hostname or IP address (required)
  - Second column 'short_name' or 'location': Custom display name (optional)
  - If no display name provided, uses full server name (not trimmed)
  - Additional columns (like 'notes') are ignored

  Additional servers are included in reports and statistics but excluded from:
  - Error counts and failure statistics
  - Email alert triggers (NTP ERROR conditions)
  - Exit code determination

  Use for monitoring servers that should not trigger alerts when failing.

Default Behavior:
  When no arguments are provided, the program will:
  - Use {ini_config['default_reference_server']} as the reference server
  - Query all A records from the {ini_config['default_discovery_domain']} domain and use them as NTP servers
  - Generate a timestamped XLSX report sorted by variance from zero
        """
    )

    # Optional arguments (made reference and servers optional for default behavior)
    parser.add_argument(
        '-r', '--reference-ntp',
        metavar='SERVER',
        help=f'Reference NTP server hostname or IP address for baseline time comparison (default: {ini_config["default_reference_server"]})'
    )

    parser.add_argument(
        '-s', '--servers-file',
        type=Path,
        metavar='FILE',
        help=f'Path to NTP server list file (.txt or .csv format) (default: query all A records from {ini_config["default_discovery_domain"]} domain)'
    )

    parser.add_argument(
        '--additional-servers',
        type=Path,
        metavar='FILE',
        help='Path to CSV file with additional NTP servers to monitor (included in report but excluded from error counts and alerts). Auto-detects "external_ntp_servers.csv" in current directory if present.'
    )

    parser.add_argument(
        '--skip-threshold',
        type=int,
        default=ini_config['skip_threshold'],
        metavar='N',
        help=f'Skip additional servers with failure_count >= this value (0 to disable skipping) [default: {ini_config["skip_threshold"]}]'
    )

    # Optional arguments with detailed help
    parser.add_argument(
        '-o', '--output-file',
        type=Path,
        metavar='FILE',
        help='Output XLSX file path (default: auto-generated with timestamp pattern "ntp_monitor_report_YYYYMMDD_HHMMSS.xlsx")'
    )

    parser.add_argument(
        '-f', '--format',
        choices=['seconds', 'milliseconds'],
        default=ini_config['default_format'],
        metavar='FORMAT',
        help=f'Delta value format: "seconds" for decimal seconds with millisecond precision, "milliseconds" for integer milliseconds [default: {ini_config["default_format"]}]'
    )

    parser.add_argument(
        '-p', '--parallel-limit',
        type=int,
        default=ini_config['default_parallel_limit'],
        metavar='N',
        help=f'Maximum number of concurrent NTP queries (1-100) [default: {ini_config["default_parallel_limit"]}]'
    )

    parser.add_argument(
        '-t', '--timeout',
        type=int,
        default=ini_config['default_timeout'],
        metavar='SECONDS',
        help=f'NTP query timeout in seconds (1-300) [default: {ini_config["default_timeout"]}]'
    )

    parser.add_argument(
        '-v', '--verbose',
        action='store_true',
        help='Enable verbose logging with detailed operation information, timing data, and error details'
    )

    parser.add_argument(
        '--version',
        action='version',
        version=f'{PROGRAM_NAME} {VERSION}'
    )

    args = parser.parse_args()

    # Set default values if not provided
    reference_ntp = args.reference_ntp or ini_config['default_reference_server']
    servers_file = args.servers_file

    # Auto-detect external_ntp_servers.csv if not specified via CLI
    additional_servers_file = args.additional_servers
    if additional_servers_file is None:
        # Look for external_ntp_servers.csv in current directory
        default_additional_file = Path('external_ntp_servers.csv')
        if default_additional_file.exists() and default_additional_file.is_file():
            additional_servers_file = default_additional_file
            logger = logging.getLogger(__name__)
            logger.debug(f"Auto-detected additional servers file: {default_additional_file}")

    # Comprehensive argument validation with detailed error messages
    validation_errors = []

    # Validate parallel limit range
    if args.parallel_limit < 1:
        validation_errors.append("Parallel limit must be at least 1")
    elif args.parallel_limit > 100:
        validation_errors.append("Parallel limit cannot exceed 100 (to prevent resource exhaustion)")

    # Validate timeout range
    if args.timeout < 1:
        validation_errors.append("Timeout must be at least 1 second")
    elif args.timeout > 300:
        validation_errors.append("Timeout cannot exceed 300 seconds (5 minutes)")

    # Validate server list file exists and is readable (only if provided)
    if servers_file:
        if not servers_file.exists():
            validation_errors.append(f"Server list file not found: {servers_file}")
        elif not servers_file.is_file():
            validation_errors.append(f"Server list path is not a file: {servers_file}")
        else:
            try:
                # Test file readability
                with open(servers_file, 'r', encoding='utf-8') as f:
                    f.read(1)  # Try to read first character
            except PermissionError:
                validation_errors.append(f"Permission denied reading server list file: {servers_file}")
            except Exception as e:
                validation_errors.append(f"Cannot read server list file {servers_file}: {e}")

    # Validate output file directory exists if specified
    if args.output_file:
        output_dir = args.output_file.parent
        if not output_dir.exists():
            validation_errors.append(f"Output directory does not exist: {output_dir}")
        elif not output_dir.is_dir():
            validation_errors.append(f"Output path parent is not a directory: {output_dir}")

    # Validate argument combinations and dependencies
    if args.format not in ['seconds', 'milliseconds']:
        validation_errors.append(f"Invalid format '{args.format}'. Must be 'seconds' or 'milliseconds'")

    # Report all validation errors at once
    if validation_errors:
        error_message = "Argument validation failed:\n" + "\n".join(f"  - {error}" for error in validation_errors)
        parser.error(error_message)

    return Config(
        reference_ntp=reference_ntp,
        ntp_servers_file=servers_file,
        additional_servers_file=additional_servers_file,
        output_file=args.output_file,
        format_type=args.format,
        parallel_limit=args.parallel_limit,
        ntp_timeout=args.timeout,
        verbose=args.verbose,
        skip_threshold=args.skip_threshold
    )


def query_reference_ntp(reference_server: str, config: Config) -> datetime:
    """
    Query primary reference NTP source at startup and return reference timestamp.

    Args:
        reference_server: Reference NTP server hostname or IP address
        config: Configuration object with timeout settings

    Returns:
        Reference timestamp in UTC for delta calculations

    Raises:
        SystemExit: If reference query fails (terminates program with error)
    """
    logger = logging.getLogger(__name__)

    try:
        logger.info(f"Querying reference NTP server: {reference_server}")

        # Query the reference NTP server
        ntp_response = query_ntp_server(reference_server, config.ntp_timeout)

        # Validate the reference response
        status, error_message = validate_ntp_response(ntp_response)

        if status != NTPStatus.OK:
            logger.error(f"Reference NTP server failed validation: {error_message}")
            logger.error("Cannot proceed without valid reference time")
            sys.exit(1)

        logger.info(f"Reference NTP query successful: {ntp_response.timestamp_utc.isoformat()}")
        logger.info(f"Reference server stratum: {ntp_response.stratum}, RTT: {ntp_response.query_rtt_ms:.2f}ms")

        return ntp_response.timestamp_utc

    except Exception as e:
        logger.error(f"Failed to query reference NTP server {reference_server}: {e}")
        logger.error("Cannot proceed without reference time - terminating")
        sys.exit(1)


def generate_default_filename() -> str:
    """
    Generate default filename with UTC timestamp pattern.

    Implements filename generation as required by:
    - Requirements 5.2: Generate filename using pattern "ntp_monitor_report_<UTCtimestamp>.xlsx"

    Returns:
        Default filename string with UTC timestamp
    """
    # Get current UTC timestamp and format for filename
    utc_now = datetime.now(timezone.utc)
    timestamp_str = utc_now.strftime("%Y%m%d_%H%M%S")

    return f"ntp_monitor_report_{timestamp_str}.xlsx"


def sort_results_by_variance(results: List[NTPResult], sort_by_variance: bool = True) -> List[NTPResult]:
    """
    Sort NTP results by variance from zero (highest to lowest absolute delta values).
    Error servers are placed at the top of the list for immediate attention.
    Additional servers are placed at the bottom as a separate section.

    Args:
        results: List of NTPResult objects to sort
        sort_by_variance: Whether to sort by variance (True) or keep original order (False)

    Returns:
        Sorted list of NTPResult objects with error servers first, then primary servers by variance, then additional servers
    """
    if not sort_by_variance:
        return results

    logger = logging.getLogger(__name__)
    logger.debug("Sorting results with error servers first, primary servers by variance, additional servers at bottom")

    # Separate primary and additional servers
    primary_results = [r for r in results if not r.is_additional_server]
    additional_results = [r for r in results if r.is_additional_server]

    # Sort primary servers: errors first, then by variance
    primary_error_results = []
    primary_results_with_delta = []
    primary_results_without_delta = []

    for result in primary_results:
        # Put error servers and Kerberos failures at the top of the list
        if result.status in [NTPStatus.ERROR, NTPStatus.TIMEOUT, NTPStatus.UNSYNCHRONIZED, NTPStatus.UNREACHABLE]:
            primary_error_results.append(result)
        elif result.kerberos_status and result.kerberos_status not in ('OK', None, ''):
            primary_error_results.append(result)
        elif result.dns_status and result.dns_status not in ('OK', None, ''):
            primary_error_results.append(result)
        elif result.ldap_status and result.ldap_status not in ('OK', None, ''):
            primary_error_results.append(result)
        elif result.ldaps_status and result.ldaps_status not in ('OK', None, ''):
            primary_error_results.append(result)
        elif result.delta_seconds is not None:
            primary_results_with_delta.append(result)
        else:
            primary_results_without_delta.append(result)

    # Sort primary error results by status severity (ERROR > TIMEOUT > UNSYNCHRONIZED > UNREACHABLE > KRB_FAIL)
    status_priority = {
        NTPStatus.ERROR: 1,
        NTPStatus.TIMEOUT: 2,
        NTPStatus.UNSYNCHRONIZED: 3,
        NTPStatus.UNREACHABLE: 4
    }
    def _error_sort_key(x):
        # NTP errors first (priority 1-4), then Kerberos failures (priority 5)
        ntp_priority = status_priority.get(x.status, 5)
        return ntp_priority

    primary_error_results.sort(key=_error_sort_key)

    # Sort primary successful results with delta by absolute value (highest variance first)
    primary_results_with_delta.sort(key=lambda x: abs(x.delta_seconds), reverse=True)

    # Sort additional servers separately: errors first, then by variance
    additional_error_results = []
    additional_results_with_delta = []
    additional_results_without_delta = []

    for result in additional_results:
        if result.status in [NTPStatus.ERROR, NTPStatus.TIMEOUT, NTPStatus.UNSYNCHRONIZED, NTPStatus.UNREACHABLE]:
            additional_error_results.append(result)
        elif result.kerberos_status and result.kerberos_status not in ('OK', None, ''):
            additional_error_results.append(result)
        elif result.dns_status and result.dns_status not in ('OK', None, ''):
            additional_error_results.append(result)
        elif result.ldap_status and result.ldap_status not in ('OK', None, ''):
            additional_error_results.append(result)
        elif result.ldaps_status and result.ldaps_status not in ('OK', None, ''):
            additional_error_results.append(result)
        elif result.delta_seconds is not None:
            additional_results_with_delta.append(result)
        else:
            additional_results_without_delta.append(result)

    # Sort additional servers the same way as primary servers
    additional_error_results.sort(key=_error_sort_key)
    additional_results_with_delta.sort(key=lambda x: abs(x.delta_seconds), reverse=True)

    # Combine: primary servers first (errors, then by variance), then additional servers at bottom
    sorted_results = (primary_error_results + primary_results_with_delta + primary_results_without_delta +
                     additional_error_results + additional_results_with_delta + additional_results_without_delta)

    logger.debug(f"Sorted {len(primary_results)} primary servers (errors first, then by variance), {len(additional_results)} additional servers at bottom")

    return sorted_results


def write_xlsx_report(results: List[NTPResult], output_path: Path, config: Config, ini_config: dict, reference_server: str) -> None:
    """
    Write NTP monitoring results to XLSX file with all required columns.

    Implements XLSX report generation with:
    - Reference server row highlighted in bold blue
    - All required columns including hostname information
    - Formatted headers with styling
    - ISO 8601 UTC timestamp formatting
    - Error status and message recording
    - Overwrite existing files by default
    - Optional sorting by variance from zero

    Args:
        results: List of NTPResult objects to write to XLSX
        output_path: Path to output XLSX file
        config: Configuration object with format settings
        ini_config: INI configuration with sorting preferences
        reference_server: Reference NTP server used for delta calculations (for highlighting)

    Raises:
        Exception: If XLSX file cannot be written
    """
    logger = logging.getLogger(__name__)

    try:
        logger.info(f"Writing XLSX report to: {output_path}")

        # Sort results by variance if configured
        sorted_results = sort_results_by_variance(results, ini_config.get('sort_by_variance', True))

        # Create a new workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "NTP Monitoring Results"

        # Define column headers with hostname information (Hostname column hidden per user request)
        headers = [
            'Timestamp (UTC)',         # Query timestamp in ISO 8601 UTC format
            'NTP Server',              # NTP server hostname/IP
            'Server IP',               # Resolved IP address
            'Short Name',              # Hostname without configured domain suffix
            'NTP Time (UTC)',          # NTP server time in ISO 8601 UTC format
            'RTT (ms)',                # Query round-trip time in milliseconds
            'Stratum',                 # NTP stratum level
            'Root Delay (ms)',         # Root delay in milliseconds
            'Root Dispersion (ms)',    # Root dispersion in milliseconds
            'Delta Value',             # Time delta in configured format
            'Delta Format',            # Format type ('seconds' or 'milliseconds')
            'Leap Indicator',          # Leap second warning (0=none, 1=+1s, 2=-1s, 3=unsync)
            'Precision',               # Clock precision as power of 2
            'Reference ID',            # Upstream time source (GPS/PPS for stratum 1, IP for 2+)
            'Reference Time (UTC)',    # When server clock was last set/corrected
            'Poll Interval (s)',       # Poll interval in seconds (2^poll)
            'Kerberos KDC',            # Kerberos KDC check result (OK, NO_RESPONSE, NOT_KDC)
            'DNS Service',             # DNS service check result (OK, NO_RESPONSE, INVALID)
            'LDAP',                    # LDAP service check result (port 389)
            'LDAPS',                   # LDAPS service check result (port 636)
            'LDAPS Cert Expiry',       # LDAPS certificate expiration date
            'Failure Count',           # Consecutive failure count for this server
            'Status',                  # Query status (OK, TIMEOUT, ERROR, etc.)
            'Error Message'            # Error details for failed queries
        ]

        # Write headers with formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows (starting from row 2)
        for row_num, result in enumerate(sorted_results, 2):  # Start from row 2 (after headers)
            # Format timestamps in ISO 8601 UTC format
            timestamp_utc_str = result.timestamp_utc.isoformat() if result.timestamp_utc else ''
            ntp_time_utc_str = result.ntp_time_utc.isoformat() if result.ntp_time_utc else ''

            # Round RTT, Root Delay, and Root Dispersion to 2 decimal places
            query_rtt_rounded = round(result.query_rtt_ms, 2) if result.query_rtt_ms is not None else ''
            root_delay_rounded = round(result.root_delay_ms, 2) if result.root_delay_ms is not None else ''
            root_dispersion_rounded = round(result.root_dispersion_ms, 2) if result.root_dispersion_ms is not None else ''

            # Prepare row data (Hostname column removed per user request)
            row_data = [
                timestamp_utc_str,
                result.ntp_server or '',
                result.ntp_server_ip or '',
                result.short_name or '',
                ntp_time_utc_str,
                query_rtt_rounded,
                result.stratum if result.stratum is not None else '',
                root_delay_rounded,
                root_dispersion_rounded,
                result.delta_formatted if result.delta_formatted is not None else '',
                config.format_type,
                result.leap_indicator if result.leap_indicator is not None else '',
                result.precision if result.precision is not None else '',
                result.ref_id or '',
                result.ref_time.isoformat() if result.ref_time else '',
                2 ** result.poll_interval if result.poll_interval is not None else '',
                result.kerberos_status or '',
                result.dns_status or '',
                result.ldap_status or '',
                result.ldaps_status or '',
                result.ldaps_cert_expiry or '',
                result.failure_count,
                result.status.value,
                result.error_message or ''
            ]

            # Write row data
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)

                # Highlight reference server row in bold blue
                if result.ntp_server == reference_server:
                    cell.font = Font(bold=True, color="0000FF")
                    logger.debug(f"Row {row_num}: Highlighted reference server '{reference_server}'")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            # Skip merged cells
            try:
                column_letter = column[0].column_letter
            except AttributeError:
                # Skip merged cells
                continue

            for cell in column:
                try:
                    # Skip merged cells
                    if hasattr(cell, 'column_letter') and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Apply status column and variance highlighting - DIRECT APPROACH ONLY
        status_col = 23  # Status column (after adding Failure Count column)
        delta_col = 10   # Delta Value column
        variance_threshold_ms = ini_config.get('variance_threshold_ms', 33)

        logger.debug(f"Applying Excel DIRECT formatting with variance threshold: {variance_threshold_ms}ms")
        highlighted_count = 0

        # Apply DIRECT cell formatting only (most reliable approach)
        # Data rows start at row 2 (row 1 = headers)
        for row_num in range(2, len(sorted_results) + 2):
            # Status column direct formatting
            status_cell = worksheet.cell(row=row_num, column=status_col)
            if status_cell.value == 'OK':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(bold=True)
            elif status_cell.value in ['ERROR', 'TIMEOUT', 'UNSYNCHRONIZED']:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(bold=True)
            elif status_cell.value == 'UNREACHABLE':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                status_cell.font = Font(bold=True)

            # Kerberos KDC column formatting (column 17)
            krb_cell = worksheet.cell(row=row_num, column=17)
            if krb_cell.value == 'OK':
                krb_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                krb_cell.font = Font(bold=True)
            elif krb_cell.value in ['NO_RESPONSE', 'KRB_ONLY', 'NOT_KDC']:
                krb_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                krb_cell.font = Font(bold=True)
            elif krb_cell.value and str(krb_cell.value).startswith('ERROR'):
                krb_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                krb_cell.font = Font(bold=True)

            # DNS Service column formatting (column 18)
            dns_cell = worksheet.cell(row=row_num, column=18)
            if dns_cell.value == 'OK':
                dns_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                dns_cell.font = Font(bold=True)
            elif dns_cell.value in ['NO_RESPONSE', 'INVALID']:
                dns_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                dns_cell.font = Font(bold=True)
            elif dns_cell.value and str(dns_cell.value).startswith('ERROR'):
                dns_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                dns_cell.font = Font(bold=True)

            # LDAP column formatting (column 19)
            ldap_cell = worksheet.cell(row=row_num, column=19)
            if ldap_cell.value == 'OK':
                ldap_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ldap_cell.font = Font(bold=True)
            elif ldap_cell.value and ldap_cell.value != '':
                ldap_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ldap_cell.font = Font(bold=True)

            # LDAPS column formatting (column 20)
            ldaps_cell = worksheet.cell(row=row_num, column=20)
            if ldaps_cell.value == 'OK':
                ldaps_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                ldaps_cell.font = Font(bold=True)
            elif ldaps_cell.value and ldaps_cell.value != '':
                ldaps_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                ldaps_cell.font = Font(bold=True)

            # Delta Value column variance highlighting - DIRECT FORMATTING ONLY
            delta_cell = worksheet.cell(row=row_num, column=delta_col)
            if delta_cell.value and isinstance(delta_cell.value, (int, float)):
                delta_value = abs(float(delta_cell.value))
                if delta_value > variance_threshold_ms:
                    try:
                        # Apply VERY STRONG direct formatting that should work everywhere
                        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        white_font = Font(color="FFFFFF", bold=True, size=12)
                        delta_cell.fill = red_fill
                        delta_cell.font = white_font

                        # Force the cell to be treated as formatted
                        delta_cell.number_format = '0'  # Ensure it's treated as a number

                        highlighted_count += 1
                        logger.info(f"Row {row_num}: DIRECT formatting applied to {delta_value}ms (exceeds {variance_threshold_ms}ms) - Cell: {delta_cell.coordinate}")
                    except Exception as e:
                        logger.error(f"Failed to apply direct formatting to row {row_num}: {e}")
                else:
                    # Apply normal formatting to non-highlighted cells
                    delta_cell.font = Font(bold=False)
                    logger.debug(f"Row {row_num}: Normal formatting {delta_value}ms (within {variance_threshold_ms}ms)")

        logger.info(f"Applied DIRECT formatting to {highlighted_count} cells exceeding {variance_threshold_ms}ms threshold")

        # Force save with specific Excel format
        try:
            # Save as Excel 2010 format (.xlsx) with explicit formatting preservation
            workbook.save(output_path)
            logger.info(f"Saved Excel file with direct formatting: {output_path}")

            # Verify the file was saved correctly by reopening it
            from openpyxl import load_workbook
            test_wb = load_workbook(output_path)
            test_ws = test_wb.active

            # Check a few cells to verify formatting was preserved
            verification_count = 0
            for row_num in range(2, min(10, len(sorted_results) + 2)):  # Check first 8 data rows (starting from row 2)
                delta_cell = test_ws.cell(row=row_num, column=delta_col)
                if delta_cell.fill and delta_cell.fill.start_color:
                    color = delta_cell.fill.start_color.rgb if hasattr(delta_cell.fill.start_color, 'rgb') else str(delta_cell.fill.start_color)
                    logger.debug(f"Row {row_num}: Cell color = {color}")
                    # Check for red color in various formats
                    if 'FF0000' in str(color) or 'ff0000' in str(color).lower():
                        verification_count += 1
                        logger.info(f"Row {row_num}: Verified red formatting - color = {color}")

            logger.info(f"Verification: Found {verification_count} cells with red formatting in saved file")
            test_wb.close()

        except Exception as e:
            logger.error(f"Error during save/verification: {e}")
            # Still try to save the file even if verification fails
            workbook.save(output_path)

        # Save the workbook
        workbook.save(output_path)

        logger.info(f"Successfully wrote {len(sorted_results)} results to XLSX file: {output_path}")

    except Exception as e:
        logger.error(f"Failed to write XLSX report to {output_path}: {e}")
        raise Exception(f"XLSX write error: {e}")


def get_output_file_path(config: Config, ini_config: dict) -> Path:
    """
    Determine output file path based on configuration.

    Implements filename generation as required by:
    - Requirements 5.2: Generate default filename with UTC timestamp pattern when not specified
    - Requirements 5.5: Handle custom output path specification
    - Uses configurable output directory from INI settings

    Args:
        config: Configuration object with output_file setting
        ini_config: INI configuration with output_directory setting

    Returns:
        Path object for output XLSX file
    """
    if config.output_file:
        # Use custom output path specification
        return config.output_file
    else:
        # Get output directory from configuration
        output_dir = Path(ini_config.get('output_directory', '.\\Reports'))

        # Create output directory if it doesn't exist
        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate default filename with UTC timestamp pattern (XLSX format)
        default_filename = generate_default_filename().replace('.csv', '.xlsx')

        # Combine directory and filename
        return output_dir / default_filename


def calculate_statistics(results: List[NTPResult]) -> SummaryStats:
    """
    Calculate min/max/avg delta values for successful measurements and count success/failure rates.
    Additional servers (is_additional_server=True) are excluded from error counts but included in delta statistics.

    Implements statistical calculation as required by:
    - Requirements 6.1: Display summary showing total servers processed and success/failure counts
    - Requirements 6.2: Calculate and display minimum, maximum, and average Delta_Value statistics

    Args:
        results: List of NTPResult objects from server processing

    Returns:
        SummaryStats object with calculated statistics (excluding additional servers from error counts)
    """
    logger = logging.getLogger(__name__)

    # Separate primary servers from additional servers for error counting
    primary_results = [r for r in results if not r.is_additional_server]

    total_servers = len(primary_results)  # Only count primary servers
    successful_servers = 0
    failed_servers = 0

    # Count servers by status type (primary servers only for error counts)
    status_counts = {}
    for status in NTPStatus:
        status_counts[status.value] = 0

    # Track if we have any error conditions for exit code determination (primary servers only)
    has_errors = False

    # Process primary servers for error counts and status tracking
    for result in primary_results:
        # Count by status
        status_counts[result.status.value] += 1

        if result.status == NTPStatus.OK:
            successful_servers += 1
        else:
            failed_servers += 1
            # Check for error conditions that should cause non-zero exit code
            if result.status in [NTPStatus.ERROR, NTPStatus.TIMEOUT, NTPStatus.UNSYNCHRONIZED]:
                has_errors = True

    # Collect delta values from ALL servers (including additional) for statistical calculations
    successful_deltas = []
    for result in results:
        if result.status == NTPStatus.OK and result.delta_seconds is not None:
            successful_deltas.append(result.delta_seconds)

    # Calculate min/max/avg delta values for successful measurements using absolute values
    min_delta = None
    max_delta = None
    avg_delta = None

    if successful_deltas:
        # Use absolute values for min/max/avg calculations
        abs_deltas = [abs(delta) for delta in successful_deltas]
        min_delta = min(abs_deltas)
        max_delta = max(abs_deltas)
        avg_delta = statistics.mean(abs_deltas)

        logger.debug(f"Delta statistics calculated from {len(successful_deltas)} successful measurements using absolute values")
        logger.debug(f"Statistics include {len([r for r in results if r.is_additional_server and r.status == NTPStatus.OK])} additional servers")
    else:
        logger.debug("No successful measurements available for delta statistics")

    return SummaryStats(
        total_servers=total_servers,
        successful_servers=successful_servers,
        failed_servers=failed_servers,
        min_delta=min_delta,
        max_delta=max_delta,
        avg_delta=avg_delta,
        status_counts=status_counts,
        has_errors=has_errors
    )


def format_summary(stats: SummaryStats, config: Config, ini_config: dict, results: List[NTPResult]) -> str:
    """
    Generate end-of-run summary text with statistics.

    Args:
        stats: SummaryStats object with calculated statistics
        config: Configuration object with format settings
        ini_config: INI configuration with threshold settings
        results: List of NTPResult objects for threshold analysis

    Returns:
        Formatted summary string for display
    """
    summary_lines = []

    # Overall processing summary
    summary_lines.append("=" * 60)
    summary_lines.append("NTP MONITORING SUMMARY")
    summary_lines.append("=" * 60)

    # Add execution environment information
    summary_lines.append(f"Hostname: {socket.gethostname()}")
    summary_lines.append(f"Execution path: {os.getcwd()}")
    summary_lines.append(f"Program: {PROGRAM_NAME} v{VERSION}")
    summary_lines.append("")

    summary_lines.append(f"Total servers processed: {stats.total_servers}")
    summary_lines.append(f"Successful queries: {stats.successful_servers}")
    summary_lines.append(f"Failed queries: {stats.failed_servers}")

    # Status breakdown
    if stats.failed_servers > 0:
        summary_lines.append("")
        summary_lines.append("Status breakdown:")
        for status_name, count in stats.status_counts.items():
            if count > 0:
                summary_lines.append(f"  {status_name}: {count}")

    # Variance threshold analysis
    variance_threshold_ms = ini_config.get('variance_threshold_ms', 33)
    exceeding_count = 0

    # Count servers exceeding variance threshold (exclude reference server and additional servers)
    reference_server = config.reference_ntp
    for result in results:
        # Exclude reference server and additional servers from threshold counting
        if (result.status == NTPStatus.OK and
            result.delta_seconds is not None and
            result.ntp_server != reference_server and
            not result.is_additional_server):
            # Use same rounding logic as Excel formatting for consistency
            delta_ms = int(round(abs(result.delta_seconds) * 1000))  # Convert to integer milliseconds like Excel
            if delta_ms > variance_threshold_ms:
                exceeding_count += 1

    summary_lines.append("")
    summary_lines.append("Variance threshold analysis:")
    summary_lines.append(f"  Threshold: {variance_threshold_ms} milliseconds")
    summary_lines.append(f"  Servers exceeding threshold: {exceeding_count}")
    if stats.successful_servers > 0:
        percentage = (exceeding_count / stats.successful_servers) * 100
        summary_lines.append(f"  Percentage exceeding: {percentage:.1f}%")

    # Delta statistics for successful measurements
    if stats.min_delta is not None and stats.max_delta is not None and stats.avg_delta is not None:
        summary_lines.append("")
        summary_lines.append("Time delta statistics (successful measurements):")

        if config.format_type == 'seconds':
            summary_lines.append(f"  Minimum delta: {stats.min_delta:.3f} seconds")
            summary_lines.append(f"  Maximum delta: {stats.max_delta:.3f} seconds")
            summary_lines.append(f"  Average delta: {stats.avg_delta:.3f} seconds")
        else:  # milliseconds
            min_ms = stats.min_delta * 1000
            max_ms = stats.max_delta * 1000
            avg_ms = stats.avg_delta * 1000
            summary_lines.append(f"  Minimum delta: {min_ms:.0f} milliseconds")
            summary_lines.append(f"  Maximum delta: {max_ms:.0f} milliseconds")
            summary_lines.append(f"  Average delta: {avg_ms:.0f} milliseconds")
    else:
        summary_lines.append("")
        summary_lines.append("No successful measurements available for delta statistics")

    summary_lines.append("=" * 60)

    return "\n".join(summary_lines)


def write_summary_file(summary_text: str, output_path: Path) -> None:
    """
    Write NTP monitoring summary to a text file.

    Args:
        summary_text: Formatted summary text to write
        output_path: Path to the XLSX output file (used to generate summary filename)
    """
    logger = logging.getLogger(__name__)

    try:
        # Generate summary filename based on XLSX filename
        summary_path = output_path.with_suffix('.txt')

        logger.info(f"Writing summary to: {summary_path}")

        # Write summary text to file
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(summary_text)
            f.write('\n')  # Ensure file ends with newline

        logger.info(f"Successfully wrote summary to: {summary_path}")

    except Exception as e:
        logger.error(f"Failed to write summary file {summary_path}: {e}")
        # Don't raise exception - summary file is not critical


def send_email_notification(summary_text: str, xlsx_path: Path, txt_path: Path, stats: SummaryStats, ini_config: dict, results: List[NTPResult], reference_server: str) -> None:
    """
    Send email notification with NTP monitoring results.

    Args:
        summary_text: Formatted summary text for email body
        xlsx_path: Path to XLSX report file for attachment
        txt_path: Path to summary text file
        stats: Summary statistics for subject line generation
        ini_config: INI configuration with email settings
        results: List of NTPResult objects for threshold analysis
        reference_server: Reference NTP server hostname for exclusion from counts
    """
    logger = logging.getLogger(__name__)

    # Check if email is enabled
    if not ini_config.get('send_email', False):
        logger.debug("Email notifications disabled in configuration")
        return

    try:
        # Determine if this is an error condition based on variance threshold
        variance_threshold_ms = ini_config.get('variance_threshold_ms', 33)
        has_error = False
        max_delta_abs = 0.0
        exceeding_count = 0

        # Filter to only primary servers (exclude additional servers) for subject line calculations
        primary_results = [r for r in results if not r.is_additional_server]

        # Count servers exceeding variance threshold (exclude reference server and additional servers)
        for result in primary_results:
            # Exclude reference server from threshold counting
            if (result.status == NTPStatus.OK and
                result.delta_seconds is not None and
                result.ntp_server != reference_server):
                # Use same rounding logic as Excel formatting for consistency
                delta_ms = int(round(abs(result.delta_seconds) * 1000))  # Convert to integer milliseconds like Excel
                if delta_ms > variance_threshold_ms:
                    exceeding_count += 1

        # Calculate max delta from primary servers only for subject line
        primary_deltas = [abs(r.delta_seconds) for r in primary_results
                         if r.status == NTPStatus.OK and r.delta_seconds is not None]

        if primary_deltas:
            max_delta_abs = max(primary_deltas)
            # Convert to milliseconds for comparison with threshold
            max_delta_ms = max_delta_abs * 1000
            has_error = max_delta_ms > variance_threshold_ms

        # Also set error flag if any servers are not responding
        if stats.failed_servers > 0:
            has_error = True

        # Check for Kerberos and DNS failures across primary servers only
        krb_fail_count = 0
        dns_fail_count = 0
        ldap_fail_count = 0
        ldaps_fail_count = 0
        for result in primary_results:
            if result.kerberos_status and result.kerberos_status not in ('OK', None, ''):
                krb_fail_count += 1
            if result.dns_status and result.dns_status not in ('OK', None, ''):
                dns_fail_count += 1
            if result.ldap_status and result.ldap_status not in ('OK', None, ''):
                ldap_fail_count += 1
            if result.ldaps_status and result.ldaps_status not in ('OK', None, ''):
                ldaps_fail_count += 1

        if krb_fail_count > 0 or dns_fail_count > 0 or ldap_fail_count > 0 or ldaps_fail_count > 0:
            has_error = True

        # Generate subject line
        domain_name = ini_config.get('default_discovery_domain', 'unknown')
        failed_count = stats.failed_servers

        if has_error:
            subject_prefix = "NTP ERROR"
        else:
            subject_prefix = "NTP REPORT"

        # Format delta section for subject - enhanced logic for threshold exceeded
        if exceeding_count > 0:
            # When servers exceed threshold, show "MAX DELTA EXCEEDED" with count
            delta_section = f"MAX DC DELTA EXCEEDED ({exceeding_count} servers)"
        else:
            # When no servers exceed threshold, show traditional "Max DC Delta" format
            if primary_deltas:
                if ini_config.get('default_format', 'seconds') == 'milliseconds':
                    max_delta_str = f"{max_delta_abs * 1000:.0f}ms"
                else:
                    max_delta_str = f"{max_delta_abs:.3f}s"
            else:
                max_delta_str = "N/A"
            delta_section = f"Max DC Delta: {max_delta_str}"

        # Build service failure details for subject
        service_issues = []
        if failed_count > 0:
            service_issues.append(f"{failed_count} NTP")
        if krb_fail_count > 0:
            service_issues.append(f"{krb_fail_count} KRB")
        if dns_fail_count > 0:
            service_issues.append(f"{dns_fail_count} DNS")
        if ldap_fail_count > 0:
            service_issues.append(f"{ldap_fail_count} LDAP")
        if ldaps_fail_count > 0:
            service_issues.append(f"{ldaps_fail_count} LDAPS")

        if service_issues:
            server_status = f"failures: {', '.join(service_issues)}"
        else:
            server_status = "all Domain Controllers responding"

        subject = f"{subject_prefix} {domain_name} - {delta_section} - {server_status}"

        logger.info(f"Sending email notification: {subject}")

        # Create email message
        msg = MIMEMultipart()
        msg['From'] = ini_config.get('from_email', 'ntp-monitor@tgna.tegna.com')
        msg['To'] = ini_config.get('to_email', 'moldham@tegna.com')
        msg['Subject'] = subject

        # Set high importance for error conditions
        if has_error:
            msg['X-Priority'] = '1'  # High priority (1=High, 3=Normal, 5=Low)
            msg['X-MSMail-Priority'] = 'High'  # Microsoft Outlook priority
            msg['Importance'] = 'High'  # Standard importance header

        # Use summary text as email body
        msg.attach(MIMEText(summary_text, 'plain'))

        # Attach XLSX file
        if xlsx_path.exists():
            with open(xlsx_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {xlsx_path.name}'
                )
                msg.attach(part)
            logger.debug(f"Attached XLSX file: {xlsx_path.name}")
        else:
            logger.warning(f"XLSX file not found for attachment: {xlsx_path}")

        # Send email
        smtp_server = ini_config.get('smtp_server', 'relay.tgna.tegna.com')
        smtp_port = ini_config.get('smtp_port', 25)
        smtp_use_tls = ini_config.get('smtp_use_tls', False)
        smtp_username = ini_config.get('smtp_username', '')
        smtp_password = ini_config.get('smtp_password', '')

        logger.debug(f"Connecting to SMTP server: {smtp_server}:{smtp_port}")

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            if smtp_use_tls:
                server.starttls()
                logger.debug("Started TLS encryption")

            if smtp_username and smtp_password:
                server.login(smtp_username, smtp_password)
                logger.debug("Authenticated with SMTP server")

            server.send_message(msg)
            logger.info(f"Email sent successfully to {msg['To']}")

    except Exception as e:
        logger.error(f"Failed to send email notification: {e}")
        # Don't raise exception - email failure shouldn't crash the program


def determine_exit_code(stats: SummaryStats) -> int:
    """
    Determine appropriate exit code based on processing results.

    Implements exit code logic as required by:
    - Requirements 6.3: Return 0 for all-success scenarios
    - Requirements 6.4: Return non-zero for any ERROR/TIMEOUT/UNSYNCHRONIZED status

    Args:
        stats: SummaryStats object with processing results

    Returns:
        Exit code: 0 for success, non-zero for failures
    """
    logger = logging.getLogger(__name__)

    if stats.has_errors:
        # Non-zero exit code for any ERROR/TIMEOUT/UNSYNCHRONIZED status
        logger.debug("Exit code: 1 (errors detected)")
        return 1
    elif stats.successful_servers == stats.total_servers:
        # All servers processed successfully
        logger.debug("Exit code: 0 (all servers successful)")
        return 0
    else:
        # Some servers failed but no critical errors
        # This handles UNREACHABLE status which is not considered a critical error
        logger.debug("Exit code: 0 (no critical errors)")
        return 0


def main():
    """
    Main entry point for NTP Delta Monitor.

    Implements main execution flow as required by:
    - Requirements 1.1, 3.1, 5.1, 6.1: Parse arguments, initialize config, execute reference query,
      process target servers concurrently, generate CSV report and summary statistics
    """
    try:
        # Load INI configuration
        ini_config = load_configuration()

        # Parse command line arguments
        config = parse_arguments()

        # Setup logging
        setup_logging(config.verbose)

        logger = logging.getLogger(__name__)
        logger.info("NTP Delta Monitor starting...")
        logger.info(f"Reference NTP: {config.reference_ntp}")

        if config.ntp_servers_file:
            logger.info(f"Server list file: {config.ntp_servers_file}")
        else:
            logger.info(f"Server discovery: Querying all A records from {ini_config['default_discovery_domain']} domain")

        logger.info(f"Parallel limit: {config.parallel_limit}")
        logger.info(f"Timeout: {config.ntp_timeout}s")
        logger.info(f"Format: {config.format_type}")

        # Initialize configuration and validate inputs
        if config.verbose:
            logger.info("Configuration validation completed successfully")

        # Validate reference server is accessible, try fallbacks if primary fails
        reference_ntp = config.reference_ntp
        logger.info(f"Validating reference NTP server: {reference_ntp}")
        reference_validated = False

        # Build list of servers to try: primary first, then fallbacks, then public NTP servers
        servers_to_try = [reference_ntp]
        for fallback in ini_config.get('fallback_servers', []):
            if fallback not in servers_to_try:
                servers_to_try.append(fallback)
        # Always include public NTP servers as last resort
        for public in ['time.cloudflare.com', 'time.google.com', 'time.aws.com', 'time.windows.com']:
            if public not in servers_to_try:
                servers_to_try.append(public)

        for candidate in servers_to_try:
            try:
                logger.info(f"Trying reference server: {candidate}")
                test_response = query_ntp_server(candidate, config.ntp_timeout)
                status, error_message = validate_ntp_response(test_response)
                if status == NTPStatus.OK:
                    reference_ntp = candidate
                    reference_validated = True
                    logger.info(f"Reference server validation successful: {candidate} (stratum {test_response.stratum})")
                    break
                else:
                    logger.warning(f"Reference server {candidate} failed validation: {error_message}")
            except Exception as e:
                logger.warning(f"Reference server {candidate} unreachable: {e}")

        if not reference_validated:
            logger.error("All reference servers failed - cannot proceed")
            sys.exit(1)

        # Update config with the working reference server
        config = Config(
            reference_ntp=reference_ntp,
            ntp_servers_file=config.ntp_servers_file,
            additional_servers_file=config.additional_servers_file,
            output_file=config.output_file,
            format_type=config.format_type,
            parallel_limit=config.parallel_limit,
            ntp_timeout=config.ntp_timeout,
            verbose=config.verbose,
            skip_threshold=config.skip_threshold
        )

        # Parse NTP server list from file or auto-discover
        if config.ntp_servers_file:
            # Use provided server list file
            ntp_servers = parse_server_file(config.ntp_servers_file)
        else:
            # Auto-discover NTP servers by getting all A records from configured domain
            logger.info(f"No server list file provided - querying all A records from {ini_config['default_discovery_domain']} domain")
            ntp_servers = discover_ntp_servers_in_domain(ini_config['default_discovery_domain'], config.ntp_timeout)

            if not ntp_servers:
                logger.warning(f"No A records found for {ini_config['default_discovery_domain']} domain, using fallback servers")
                # Use fallback servers from INI configuration
                ntp_servers = ini_config['fallback_servers']
                logger.info(f"Using fallback server list: {', '.join(ntp_servers)}")

        if not ntp_servers:
            logger.error("No NTP servers available for monitoring")
            return 1

        logger.info(f"Loaded {len(ntp_servers)} NTP servers for monitoring")

        # Process target NTP servers concurrently
        logger.info("Processing target NTP servers...")
        results = process_servers_parallel(ntp_servers, config.reference_ntp, config, ini_config)

        # Process additional servers if specified (excluded from error counts)
        additional_servers_with_counts = []  # Store for CSV write-back
        if config.additional_servers_file:
            logger.info(f"Processing additional NTP servers from: {config.additional_servers_file}")
            try:
                # Use specialized parser for additional servers CSV with all 6 tracking fields
                additional_servers_with_counts = parse_additional_servers_csv(config.additional_servers_file)
                if additional_servers_with_counts:
                    logger.info(f"Loaded {len(additional_servers_with_counts)} additional NTP servers for monitoring")

                    # Filter out servers that have exceeded the skip threshold
                    if config.skip_threshold > 0:
                        active_servers = []
                        skipped_servers = []
                        for entry in additional_servers_with_counts:
                            server, short_name, failure_count = entry[0], entry[1], entry[2]
                            if failure_count >= config.skip_threshold:
                                skipped_servers.append(entry)
                                logger.info(f"Skipping server {server} ({short_name}): failure_count={failure_count} >= threshold={config.skip_threshold}")
                            else:
                                active_servers.append(entry)
                        if skipped_servers:
                            logger.info(f"Skipped {len(skipped_servers)} servers exceeding failure threshold ({config.skip_threshold})")
                        servers_to_process = active_servers
                    else:
                        servers_to_process = additional_servers_with_counts

                    # Create tracking dictionaries for active servers only
                    failure_counts = {server: failure_count for server, _, failure_count, _, _, _, _ in servers_to_process}
                    missed_counts = {server: missed for server, _, _, missed, _, _, _ in servers_to_process}
                    failed_statuses = {server: failed for server, _, _, _, failed, _, _ in servers_to_process}
                    time_counts = {server: time for server, _, _, _, _, _, time in servers_to_process}
                    logger.debug(f"Initialized tracking for {len(failure_counts)} servers")

                    # Process servers (pass tuples of server, short_name for processing)
                    additional_servers_for_processing = [(server, short_name) for server, short_name, _, _, _, _, _ in servers_to_process]
                    additional_results = process_servers_parallel(additional_servers_for_processing, config.reference_ntp, config, ini_config, is_additional_servers=True, failure_counts=failure_counts, missed_counts=missed_counts, failed_statuses=failed_statuses, time_counts=time_counts)
                    results.extend(additional_results)
                    logger.info(f"Combined results: {len(results)} total servers ({len(ntp_servers)} primary + {len(servers_to_process)} additional, {len(additional_servers_with_counts) - len(servers_to_process)} skipped)")
                else:
                    logger.warning(f"No additional servers found in {config.additional_servers_file}")
            except Exception as e:
                logger.error(f"Failed to process additional servers file {config.additional_servers_file}: {e}")
                logger.info("Continuing with primary servers only")

        # Generate XLSX report
        output_path = get_output_file_path(config, ini_config)
        write_xlsx_report(results, output_path, config, ini_config, config.reference_ntp)

        # Generate summary statistics
        stats = calculate_statistics(results)
        summary_text = format_summary(stats, config, ini_config, results)

        # Write summary to text file
        write_summary_file(summary_text, output_path)

        # Send email notification if enabled
        txt_path = output_path.with_suffix('.txt')
        send_email_notification(summary_text, output_path, txt_path, stats, ini_config, results, config.reference_ntp)

        # Write updated tracking counters back to CSV (if additional servers were processed)
        if config.additional_servers_file and additional_servers_with_counts:
            logger.info("Writing updated tracking counters to CSV")
            try:
                # Update all tracking fields from results
                updated_servers = []
                for server, short_name, old_failure_count, old_missed, old_failed, old_dlist, old_time in additional_servers_with_counts:
                    # Find the result for this server
                    result = next((r for r in results if r.ntp_server == server and r.is_additional_server), None)
                    if result:
                        # Use all updated tracking fields from the result
                        updated_servers.append((server, short_name, result.failure_count, result.missed, result.failed, old_dlist, result.time))
                        logger.debug(f"Server {server}: failure_count={result.failure_count}, missed={result.missed}, failed={result.failed}, time={result.time}")
                    else:
                        # Server wasn't processed (shouldn't happen), keep old values
                        updated_servers.append((server, short_name, old_failure_count, old_missed, old_failed, old_dlist, old_time))
                        logger.warning(f"Server {server}: no result found, keeping old values")

                # Write updated data to CSV
                write_additional_servers_csv(config.additional_servers_file, updated_servers)
                logger.info(f"Successfully updated tracking counters in {config.additional_servers_file}")
            except Exception as e:
                logger.error(f"Failed to write updated tracking counters: {e}")
                logger.debug("Continuing despite CSV write failure")

        # Display end-of-run summary
        print(summary_text)

        # Determine appropriate exit code
        exit_code = determine_exit_code(stats)

        if exit_code == 0:
            logger.info("NTP monitoring completed successfully")
        else:
            logger.warning("NTP monitoring completed with errors")

        return exit_code

    except KeyboardInterrupt:
        print("\nOperation cancelled by user", file=sys.stderr)
        return 1
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
